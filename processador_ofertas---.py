# -*- coding: utf-8 -*-
"""
Script para processamento de arquivos XML de ofertas (processador_ofertas.py).
V9.3: Reintroduzida e aplicada a COLUMNS_FORMATTING_CONFIG para formatação
      detalhada de colunas no relatório 'giro_da_praça_ofertas.xlsx'.
      A formatação de destaque (modo update) tem precedência para fundo/fonte.
V9.5 (Solicitado): Alterada a organização do relatório 'giro_da_praça_ofertas.xlsx'
      para NOME_PROMOÇÃO -> SESSÃO -> PRODUTO, e reintroduzido espaçamento
      entre grupos de seção dentro da mesma promoção.
"""

# ------------------------------------------
# DEPENDÊNCIAS
# ------------------------------------------
import pandas as pd
import os
import glob
import xmltodict # Certifique-se de que está instalado: pip install xmltodict
from typing import List, Dict, Any, Optional
import re
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import traceback
import argparse # Para argumentos CLI

# ------------------------------------------
# CONSTANTES
# ------------------------------------------
INPUT_XML_DIR = './arquivos_xml_entrada/'
XML_EXTENSION = '.xml'
PRODUCT_MASTER_FILE_NAME = 'produtos_cadastrados.xlsx'
FINAL_OFFERS_REPORT_NAME = 'giro_da_praça_ofertas.xlsx'

SECOES_EXCECAO_AGRUPAMENTO = ["#01 MERCEARIA - #01 ALTO GIRO", "#01 MERCEARIA - #02 ALTO GIRO"]

# Estilos para destaque de linha no modo update
HIGHLIGHT_FILL = PatternFill(start_color="45A045", end_color="45A045", fill_type="solid") # Verde
HIGHLIGHT_FONT = Font(color="FFFFFF", bold=True) # Fonte Branca, Negrito

# Configuração de Formatação de Colunas para giro_da_praça_ofertas.xlsx
COLUMNS_FORMATTING_CONFIG = {
  '_DEFAULT_': {
      'font_name': 'Calibri',
      'font_size': 10,
      'font_color': "000000",
      'font_bold': False,
      'font_italic': False,
      'number_format': '@',
      'alignment_horizontal': 'left',
      'alignment_vertical': 'center',
      'fill_color': None, # Sem preenchimento por padrão para dados

      'header_font_name': 'Calibri',
      'header_font_size': 9,
      'header_font_color': "000000",
      'header_font_bold': True,
      'header_font_italic': False,
      'header_fill_color': "FFFFFF", # Fundo branco para cabeçalho por padrão
      'header_alignment_horizontal': 'left',
      'header_alignment_vertical': 'center',
  },
  'NOME_PROMOÇÃO': {
      'font_size': 7,
  },
  'SESSÃO': {
      'font_size': 7,
      'font_color': "000000",
  },
  'ID': {
      'font_size': 7 ,
      'number_format': '0',
  },
  'PRODUTO': {
      'font_size': 9,
  },
  'TIPO': {
      'font_size': 9,
      'font_color': "000000",
      'font_italic': True
  },
  'PROMOÇÃO': {
      'font_size': 10,
      'font_color': "000000",
      'font_bold': True,
      'number_format': 'R$ #,##0.00',
      'alignment_horizontal': 'right',
  }
}
# ------------------------------------------
# FUNÇÕES PARA TRATAMENTO DE TEXTO
# ------------------------------------------

def remover_acentuacao(texto: str) -> str:
    if not isinstance(texto, str): return ""
    mapeamento_acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u', 'ã': 'a', 'õ': 'o', 'ä': 'a', 'ë': 'e', 'ï': 'i',
        'ö': 'o', 'ü': 'u', 'ç': 'c', 'ñ': 'n', 'ÿ': 'y', 'ý': 'y', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O',
        'Ú': 'U', 'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U', 'Â': 'A', 'Ê': 'E', 'Î': 'I', 'Ô': 'O',
        'Û': 'U', 'Ã': 'A', 'Õ': 'O', 'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U', 'Ç': 'C', 'Ñ': 'N',
        'Ÿ': 'Y', 'Ý': 'Y'
    }
    return ''.join(mapeamento_acentos.get(c, c) for c in texto)

def normalizar_texto(texto: str) -> str:
    if not isinstance(texto, str): return ""
    texto_sem_acentos = remover_acentuacao(texto)
    return texto_sem_acentos.lower()

# ------------------------------------------
# FUNÇÕES PARA MANIPULAÇÃO DE ARQUIVOS
# ------------------------------------------

def listar_arquivos_xml(diretorio: str) -> List[str]:
    return glob.glob(os.path.join(diretorio, f'*{XML_EXTENSION}'))

def renomear_arquivo(arquivo_antigo: str, diretorio: str) -> None:
    nome_arquivo = os.path.basename(arquivo_antigo)
    novo_nome = normalizar_texto(nome_arquivo)
    novo_caminho = os.path.join(diretorio, novo_nome)
    try:
        if arquivo_antigo != novo_caminho:
            os.rename(arquivo_antigo, novo_caminho)
    except FileExistsError:
        print(f"⚠️  Arquivo '{novo_caminho}' já existe. Não foi renomeado '{arquivo_antigo}'.")
    except Exception as e:
        print(f"❌ Erro ao renomear '{arquivo_antigo}' para '{novo_caminho}': {e}")

def renomear_arquivos_em_lote(arquivos: List[str], diretorio: str) -> None:
    for arquivo in arquivos:
        renomear_arquivo(arquivo, diretorio)

# ------------------------------------------
# FUNÇÕES PARA PROCESSAMENTO DE XML
# ------------------------------------------

def parse_produto_xml(produto_xml_data: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "Promoção_XML": produto_xml_data.get('descrpromocao', ''),
        "ID_Produto_XML": int(produto_xml_data.get('idsubproduto', 0)),
        "Nome_Produto_XML": produto_xml_data.get('descrresproduto', ''),
        "Preco_Promocao_XML": float(produto_xml_data.get('precopromocao', 0.0) if produto_xml_data.get('precopromocao') not in [None, ''] else 0.0)
    }

def processar_arquivo_xml(nome_arquivo_xml: str) -> List[Dict[str, Any]]:
    caminho_arquivo = os.path.join(INPUT_XML_DIR, nome_arquivo_xml)
    try:
        with open(caminho_arquivo, 'rb') as arquivo_xml:
            dados_xml = xmltodict.parse(arquivo_xml)
            if 'temporario_846' not in dados_xml:
                print(f"⚠️ Estrutura 'temporario_846' não encontrada em {nome_arquivo_xml}.")
                return []
            temporario_846 = dados_xml['temporario_846']
            produtos_data_xml = temporario_846.get('temporario_846_row', [])
            if isinstance(produtos_data_xml, dict):
                produtos_data_xml = [produtos_data_xml]
            if not produtos_data_xml:
                return []
            return [parse_produto_xml(p) for p in produtos_data_xml if isinstance(p, dict)]
    except FileNotFoundError:
        print(f"❌ Arquivo XML não encontrado: {caminho_arquivo}")
        return []
    except Exception as e:
        print(f"❌ Erro ao processar XML {nome_arquivo_xml}: {e}")
        traceback.print_exc()
        return []

# ------------------------------------------
# FUNÇÃO PARA CONSOLIDAÇÃO DE DADOS XML
# ------------------------------------------

def consolidar_dados_de_xmls(lista_arquivos_xml: List[str]) -> pd.DataFrame:
    todos_produtos_lista = []
    for arquivo_nome in lista_arquivos_xml:
        produtos_do_arquivo = processar_arquivo_xml(arquivo_nome)
        if produtos_do_arquivo:
            todos_produtos_lista.extend(produtos_do_arquivo)
            print(f"✅ {arquivo_nome}: {len(produtos_do_arquivo)} produtos processados.")
    if not todos_produtos_lista:
        print("⚠️ Nenhum produto encontrado nos arquivos XML.")
        return pd.DataFrame()
    return pd.DataFrame(todos_produtos_lista)

# ------------------------------------------
# FUNÇÕES AUXILIARES PARA INTEGRAÇÃO DE DADOS
# ------------------------------------------

def extract_product_base_name(product_name: str) -> str:
    if not isinstance(product_name, str): return ""
    unidades = [
        "KILOGRAMAS", "KILOGRAMA", "QUILOGRAMAS", "QUILOGRAMA", "LITROS", "LITRO", "UNIDADES", "UNIDADE",
        "PACOTES", "PACOTE", "CAIXAS", "CAIXA", "FARDOS", "FARDO", "GARRAFAS", "GARRAFA", "LATAS", "LATA",
        "POTES", "POTE", "ROLOS", "ROLO", "SACHES", "SACHE", "TUBOS", "TUBO", "VIDROS", "VIDRO", "SACOLAS",
        "SACOLA", "BLISTERS", "BLISTER", "CARTELAS", "CARTELA", "DISPLAYS", "DISPLAY", "GALÕES", "GALÃO",
        "REFILs", "REFIL", "TABLETES", "TABLETE", "PEÇAS", "PEÇA", "FOLHAS", "FOLHA", "FLACONETES",
        "FLACONETE", "BISNAGAS", "BISNAGA", "AMPOLAS", "AMPOLA", "DRAGEAS", "CAPSULAS", "CAPSULA",
        "COMPRIMIDOS", "COMPRIMIDO", "ENVELOPES", "ENVELOPE",
        "KG", "GR", "G", "MG", "L", "ML", "CM", "MM", "M",
        "UN", "UND", "UNID", "PC", "PÇ", "PCT", "CX", "FD", "GF", "LT", "PT", "RL", "SC", "TB", "VD",
        "BL", "CT", "DP", "GL", "RF", "TBL", "FL", "CAPS", "COMP", "ENV"
    ]
    product_name_norm = ' '.join(product_name.split())
    unidades_regex_part = "|".join(map(re.escape, sorted(unidades, key=len, reverse=True)))
    regex_str = r"^(.*?)(\s?)(\d+[\.,]?\d*)(\s*)(" + unidades_regex_part + r")($|\s+.*$)"
    match = re.match(regex_str, product_name_norm, re.IGNORECASE)

    if match:
        nome_antes_da_medida = match.group(1).strip()
        quantidade = match.group(3)
        unidade_texto = match.group(5)
        base_nome = f"{nome_antes_da_medida} {quantidade}{unidade_texto}".strip()
        return ' '.join(base_nome.split())
    return product_name_norm

def extract_product_type_from_full_name(full_name: str, base_name: str) -> str:
    if not isinstance(full_name, str) or not isinstance(base_name, str): return ""
    full_name_norm = ' '.join(full_name.split())
    base_name_norm = ' '.join(base_name.split())
    if full_name_norm == base_name_norm: return ""
    if full_name_norm.startswith(base_name_norm):
        type_part = full_name_norm[len(base_name_norm):].strip()
        type_part = re.sub(r"^[ \-\/\_]+", "", type_part).strip()
        return type_part
    return ""

def get_section_sort_key(section_name: Any) -> str:
    if not isinstance(section_name, str) or pd.isna(section_name): return "SEM SEÇÃO DEFINIDA"
    if section_name.startswith("#01 MERCEARIA - #") and "ALTO GIRO" in section_name:
        return "#01 MERCEARIA - ALTO GIRO"
    return section_name.split(" - ", 1)[0]

# ------------------------------------------
# FUNÇÕES PARA INTEGRAÇÃO E MANUTENÇÃO DE DADOS
# ------------------------------------------

def atualizar_cadastro_produtos_base(df_ofertas_xml: pd.DataFrame) -> None:
    path_cadastro_base = f'./{PRODUCT_MASTER_FILE_NAME}'
    cols_cadastro_base_ordenadas = ['ID', 'SESSÃO', 'NOME_SISTEMA', 'NOME_CORRIGIDO']

    try:
        df_cadastro_base = pd.read_excel(path_cadastro_base)
        if 'ID' not in df_cadastro_base.columns:
            print(f"⚠️  Arquivo '{PRODUCT_MASTER_FILE_NAME}' não possui coluna 'ID'. Tratando como novo.")
            df_cadastro_base = pd.DataFrame(columns=cols_cadastro_base_ordenadas)
        temp_df = pd.DataFrame(columns=cols_cadastro_base_ordenadas)
        for col in cols_cadastro_base_ordenadas:
            if col in df_cadastro_base.columns:
                temp_df[col] = df_cadastro_base[col]
            else:
                temp_df[col] = np.nan
        df_cadastro_base = temp_df
    except FileNotFoundError:
        print(f"ℹ️  Arquivo '{PRODUCT_MASTER_FILE_NAME}' não encontrado. Será criado.")
        df_cadastro_base = pd.DataFrame(columns=cols_cadastro_base_ordenadas)
    except Exception as e:
        print(f"❌ Erro ao carregar '{PRODUCT_MASTER_FILE_NAME}': {e}. Operação cancelada.")
        return

    novos_produtos_df = pd.DataFrame()

    if df_ofertas_xml.empty:
        print(f"ℹ️ DataFrame de ofertas XML está vazio. Nenhum produto novo para adicionar ao '{PRODUCT_MASTER_FILE_NAME}'.")
    else:
        if 'ID_Produto_XML' not in df_ofertas_xml.columns or 'Nome_Produto_XML' not in df_ofertas_xml.columns:
            print(f"❌ Colunas 'ID_Produto_XML' ou 'Nome_Produto_XML' ausentes no DataFrame de ofertas. Não é possível atualizar o cadastro.")
        else:
            if 'ID' in df_cadastro_base.columns:
                df_cadastro_base['ID'] = pd.to_numeric(df_cadastro_base['ID'], errors='coerce')
                ids_existentes_cadastro = set(df_cadastro_base['ID'].dropna().astype(int).unique())
            else:
                ids_existentes_cadastro = set()

            df_ofertas_xml['ID_Produto_XML'] = pd.to_numeric(df_ofertas_xml['ID_Produto_XML'], errors='coerce')
            novos_produtos_df = df_ofertas_xml[
                ~df_ofertas_xml['ID_Produto_XML'].dropna().astype(int).isin(ids_existentes_cadastro)
            ].copy()

            if novos_produtos_df.empty:
                print(f"ℹ️ Nenhum produto novo das ofertas XML para adicionar ao '{PRODUCT_MASTER_FILE_NAME}'.")
            else:
                print(f"ℹ️ {len(novos_produtos_df)} novos produtos encontrados para adicionar.")
                novos_para_cadastro_list = []
                for _, row in novos_produtos_df.iterrows():
                    novos_para_cadastro_list.append({
                        'ID': row['ID_Produto_XML'],
                        'NOME_SISTEMA': row['Nome_Produto_XML'],
                        'SESSÃO': 'SEÇÃO NÃO ESPECIFICADA',
                        'NOME_CORRIGIDO': row['Nome_Produto_XML']
                    })
                df_novos_formatados = pd.DataFrame(novos_para_cadastro_list, columns=cols_cadastro_base_ordenadas)
                df_cadastro_base = pd.concat([df_cadastro_base, df_novos_formatados], ignore_index=True)
                df_cadastro_base.drop_duplicates(subset=['ID'], keep='first', inplace=True)

    try:
        df_para_salvar = df_cadastro_base.reindex(columns=cols_cadastro_base_ordenadas)
        df_para_salvar = df_para_salvar.sort_values(by=['SESSÃO', 'NOME_SISTEMA', 'ID'])


        if 'ID' in df_para_salvar.columns:
            df_para_salvar['ID'] = pd.to_numeric(df_para_salvar['ID'], errors='coerce')

        df_para_salvar.to_excel(path_cadastro_base, index=False)

        if not df_para_salvar.empty or os.path.exists(path_cadastro_base):
            no_fill = PatternFill(fill_type=None)

            ids_produtos_oferta_xml = set()
            if not df_ofertas_xml.empty and 'ID_Produto_XML' in df_ofertas_xml.columns:
                ids_produtos_oferta_xml = set(df_ofertas_xml['ID_Produto_XML'].dropna().astype(int).unique())

            produtos_coloridos_info_list = []

            wb = load_workbook(path_cadastro_base)
            ws = wb.active

            id_col_index = None
            nome_corrigido_col_index = None
            nome_sistema_col_index = None

            if ws.max_row > 0:
                header_cells = ws[1]
                for i, cell in enumerate(header_cells):
                    if cell.value == 'ID': id_col_index = i + 1
                    elif cell.value == 'NOME_CORRIGIDO': nome_corrigido_col_index = i + 1
                    elif cell.value == 'NOME_SISTEMA': nome_sistema_col_index = i + 1

            if id_col_index is None and ws.max_row > 0 :
                print(f"⚠️  Aviso: Coluna 'ID' não encontrada no cabeçalho de '{PRODUCT_MASTER_FILE_NAME}'. Não será possível aplicar cores.")
            elif ws.max_row <= 1 and not ids_produtos_oferta_xml:
                 print(f"ℹ️  Arquivo '{PRODUCT_MASTER_FILE_NAME}' está vazio ou contém apenas cabeçalho e não há produtos XML para colorir.")
            elif ws.max_row > 1 and id_col_index is not None:
                for row_num in range(2, ws.max_row + 1):
                    for col_num in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col_num).fill = no_fill

                    id_cell_value = ws.cell(row=row_num, column=id_col_index).value
                    if id_cell_value is not None:
                        try:
                            current_id = int(float(id_cell_value))
                            if current_id in ids_produtos_oferta_xml:
                                for col_num in range(1, ws.max_column + 1):
                                    ws.cell(row=row_num, column=col_num).fill = HIGHLIGHT_FILL

                                nome_produto = "NOME NÃO ENCONTRADO"
                                if nome_corrigido_col_index and ws.cell(row=row_num, column=nome_corrigido_col_index).value:
                                    nome_produto = ws.cell(row=row_num, column=nome_corrigido_col_index).value
                                elif nome_sistema_col_index and ws.cell(row=row_num, column=nome_sistema_col_index).value:
                                     nome_produto = ws.cell(row=row_num, column=nome_sistema_col_index).value
                                produtos_coloridos_info_list.append({'ID': current_id, 'Nome': nome_produto})
                        except (ValueError, TypeError): pass

            wb.save(path_cadastro_base)

            if not novos_produtos_df.empty:
                print(f"✅ '{PRODUCT_MASTER_FILE_NAME}' atualizado com {len(novos_produtos_df)} novos produtos e formatação de cores aplicada.")
            else:
                print(f"✅ '{PRODUCT_MASTER_FILE_NAME}' salvo com formatação de cores aplicada.")

            if produtos_coloridos_info_list:
                print(f"ℹ️ Produtos das ofertas XML (coloridos em verde em '{PRODUCT_MASTER_FILE_NAME}'):")
                unique_produtos_coloridos = [dict(t) for t in {tuple(d.items()) for d in produtos_coloridos_info_list}]
                for prod_info in sorted(unique_produtos_coloridos, key=lambda x: x['ID']):
                    print(f"  - ID: {prod_info['ID']}, Nome: {str(prod_info['Nome'])}")
            elif ids_produtos_oferta_xml:
                print(f"ℹ️ Havia {len(ids_produtos_oferta_xml)} produtos nas ofertas XML, mas nenhum correspondente foi encontrado ou pode ser colorido em '{PRODUCT_MASTER_FILE_NAME}'.")
            else:
                print(f"ℹ️ Nenhum produto das ofertas XML para colorir em '{PRODUCT_MASTER_FILE_NAME}'.")

    except Exception as e:
        print(f"❌ Erro ao salvar ou formatar '{PRODUCT_MASTER_FILE_NAME}': {e}")
        traceback.print_exc()

def gerar_relatorio_ofertas_finais(df_ofertas_xml: pd.DataFrame, mode: str = 'create') -> None:
    path_cadastro_base = f'./{PRODUCT_MASTER_FILE_NAME}'
    path_relatorio_final = f'./{FINAL_OFFERS_REPORT_NAME}'
    colunas_finais_excel_ordenadas = ['NOME_PROMOÇÃO', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO']
    colunas_com_destaque = colunas_finais_excel_ordenadas + ['DESTAQUE', 'Chave_Grupo_Secao_Ordenacao'] # Adicionar Chave aqui para DataFrame interno

    try:
        df_cadastro_base = pd.read_excel(path_cadastro_base)
        colunas_cadastro_esperadas = ['ID', 'SESSÃO', 'NOME_SISTEMA', 'NOME_CORRIGIDO']
        if not all(col in df_cadastro_base.columns for col in colunas_cadastro_esperadas):
            missing_cols = [col for col in colunas_cadastro_esperadas if col not in df_cadastro_base.columns]
            print(f"❌ Arquivo '{PRODUCT_MASTER_FILE_NAME}' não contém todas as colunas esperadas. Faltando: {missing_cols}. Não é possível gerar relatório.")
            return
    except FileNotFoundError:
        print(f"❌ Arquivo de cadastro '{PRODUCT_MASTER_FILE_NAME}' não encontrado. Execute a atualização do cadastro primeiro."); return
    except Exception as e:
        print(f"❌ Erro ao carregar '{PRODUCT_MASTER_FILE_NAME}': {e}"); return

    if df_ofertas_xml.empty and mode == 'create':
        print(f"⚠️ DataFrame de ofertas XML está vazio. Nada a processar para '{FINAL_OFFERS_REPORT_NAME}' no modo 'create'."); return
    elif df_ofertas_xml.empty and mode == 'update':
        print(f"⚠️ DataFrame de ofertas XML está vazio. Tentando gerar relatório '{FINAL_OFFERS_REPORT_NAME}' para limpar/atualizar com base no anterior.")

    required_xml_cols = ['ID_Produto_XML', 'Promoção_XML', 'Preco_Promocao_XML', 'Nome_Produto_XML']
    if not df_ofertas_xml.empty and not all(col in df_ofertas_xml.columns for col in required_xml_cols):
        print(f"❌ DataFrame de ofertas XML não contém todas as colunas requeridas: {required_xml_cols}.")
        return

    df_ofertas_xml_copy = df_ofertas_xml[required_xml_cols].copy() if not df_ofertas_xml.empty else pd.DataFrame(columns=required_xml_cols)
    df_cadastro_base_copy = df_cadastro_base[['ID', 'SESSÃO', 'NOME_CORRIGIDO']].copy()

    df_ofertas_xml_copy['ID_Produto_XML'] = pd.to_numeric(df_ofertas_xml_copy['ID_Produto_XML'], errors='coerce').astype('Int64')
    df_cadastro_base_copy['ID'] = pd.to_numeric(df_cadastro_base_copy['ID'], errors='coerce').astype('Int64')

    if not df_ofertas_xml_copy.empty:
        df_merged = pd.merge(
            df_ofertas_xml_copy,
            df_cadastro_base_copy,
            left_on='ID_Produto_XML',
            right_on='ID',
            how='left'
        )

        df_merged = df_merged.rename(columns={
            'Promoção_XML': 'NOME_PROMOÇÃO_ORIGINAL',
            'Preco_Promocao_XML': 'PROMOÇÃO',
            'ID_Produto_XML': 'ID_FINAL'
        })
        if 'ID' in df_merged.columns and 'ID_FINAL' in df_merged.columns:
            df_merged['ID'] = df_merged['ID'].fillna(df_merged['ID_FINAL'])
        elif 'ID_FINAL' in df_merged.columns :
            df_merged.rename(columns={'ID_FINAL': 'ID'}, inplace=True)

        df_merged['NOME_CORRIGIDO'] = df_merged['NOME_CORRIGIDO'].fillna(df_merged['Nome_Produto_XML'])
        df_merged['SESSÃO'] = df_merged['SESSÃO'].fillna('SEÇÃO NÃO ESPECIFICADA')
        df_merged['NOME_PROMOÇÃO_ORIGINAL'] = df_merged['NOME_PROMOÇÃO_ORIGINAL'].fillna('')

        df_merged['Produto_Base_Agrupamento'] = df_merged['NOME_CORRIGIDO'].apply(lambda x: extract_product_base_name(str(x)))
        df_merged['Tipo_Produto_Variante'] = df_merged.apply(lambda row: extract_product_type_from_full_name(str(row['NOME_CORRIGIDO']), str(row['Produto_Base_Agrupamento'])), axis=1)
        df_merged['Excecao_Agrupamento'] = df_merged['SESSÃO'].isin(SECOES_EXCECAO_AGRUPAMENTO)
    else:
        expected_merged_cols = ['ID', 'NOME_PROMOÇÃO_ORIGINAL', 'PROMOÇÃO', 'Nome_Produto_XML',
                                'SESSÃO', 'NOME_CORRIGIDO', 'Produto_Base_Agrupamento',
                                'Tipo_Produto_Variante', 'Excecao_Agrupamento']
        df_merged = pd.DataFrame(columns=expected_merged_cols)

    lista_dfs_processados = []

    if not df_merged.empty:
        df_excecoes_diretas = df_merged[df_merged['Excecao_Agrupamento']].copy()
        if not df_excecoes_diretas.empty:
            df_temp_data = {
                'NOME_PROMOÇÃO': df_excecoes_diretas['NOME_PROMOÇÃO_ORIGINAL'],
                'SESSÃO': df_excecoes_diretas['SESSÃO'],
                'ID': df_excecoes_diretas['ID'],
                'PRODUTO': df_excecoes_diretas['NOME_CORRIGIDO'],
                'TIPO': "",
                'PROMOÇÃO': df_excecoes_diretas['PROMOÇÃO'],
                'DESTAQUE': False
                # Chave_Grupo_Secao_Ordenacao será adicionada depois globalmente a df_relatorio_parcial
            }
            df_temp = pd.DataFrame(df_temp_data)
            lista_dfs_processados.append(df_temp) # Não reindexar aqui para não perder colunas que podem não estar em colunas_com_destaque ainda

        df_potenciais = df_merged[~df_merged['Excecao_Agrupamento']].copy()
        if not df_potenciais.empty:
            df_potenciais['Contagem_No_Grupo'] = df_potenciais.groupby(
                ['Produto_Base_Agrupamento', 'PROMOÇÃO']
            )['ID'].transform('size')

            df_para_agregar = df_potenciais[df_potenciais['Contagem_No_Grupo'] > 1].copy()
            if not df_para_agregar.empty:
                def aggregate_types(series_types):
                    valid_types = [str(t).strip() for t in series_types if pd.notna(t) and str(t).strip()]
                    return ', '.join(sorted(list(set(valid_types)))) if valid_types else ""

                df_agregado = df_para_agregar.groupby(
                    ['Produto_Base_Agrupamento', 'PROMOÇÃO'], as_index=False
                ).agg(
                    TIPO=pd.NamedAgg(column='Tipo_Produto_Variante', aggfunc=aggregate_types),
                    NOME_PROMOÇÃO=pd.NamedAgg(column='NOME_PROMOÇÃO_ORIGINAL', aggfunc='first'),
                    ID=pd.NamedAgg(column='ID', aggfunc='first'),
                    SESSÃO=pd.NamedAgg(column='SESSÃO', aggfunc='first')
                ).rename(columns={'Produto_Base_Agrupamento': 'PRODUTO'})

                df_agregado['DESTAQUE'] = False
                lista_dfs_processados.append(df_agregado)

            df_unicos_em_grupo = df_potenciais[df_potenciais['Contagem_No_Grupo'] == 1].copy()
            if not df_unicos_em_grupo.empty:
                df_temp_data = {
                    'NOME_PROMOÇÃO': df_unicos_em_grupo['NOME_PROMOÇÃO_ORIGINAL'],
                    'SESSÃO': df_unicos_em_grupo['SESSÃO'],
                    'ID': df_unicos_em_grupo['ID'],
                    'PRODUTO': df_unicos_em_grupo['NOME_CORRIGIDO'],
                    'TIPO': "",
                    'PROMOÇÃO': df_unicos_em_grupo['PROMOÇÃO'],
                    'DESTAQUE': False
                }
                df_temp = pd.DataFrame(df_temp_data)
                lista_dfs_processados.append(df_temp)

    if lista_dfs_processados:
        df_relatorio_parcial = pd.concat(lista_dfs_processados, ignore_index=True)
    else:
        # Criar com as colunas esperadas para o processamento, incluindo as auxiliares se necessário
        df_relatorio_parcial = pd.DataFrame(columns=[c for c in colunas_com_destaque if c != 'Chave_Grupo_Secao_Ordenacao'])


    # Garantir colunas essenciais e a coluna DESTAQUE
    for col in colunas_finais_excel_ordenadas: # Garante colunas que vão para o Excel
        if col not in df_relatorio_parcial.columns:
            df_relatorio_parcial[col] = np.nan if col in ['ID', 'PROMOÇÃO'] else ""

    if 'DESTAQUE' not in df_relatorio_parcial.columns:
        df_relatorio_parcial['DESTAQUE'] = False
    df_relatorio_parcial['DESTAQUE'] = df_relatorio_parcial['DESTAQUE'].fillna(False).astype(bool)


    df_antigo_relatorio = None
    chaves_antigas_presentes = set()
    mapa_precos_antigos = {}
    if mode == 'update':
        try:
            df_antigo_relatorio = pd.read_excel(
                path_relatorio_final,
                sheet_name='Ofertas Finais',
                dtype={'PRODUTO': str, 'TIPO': str, 'SESSÃO': str, 'NOME_PROMOÇÃO': str, 'ID': 'Int64'}
            )
            df_antigo_relatorio['PRODUTO'] = df_antigo_relatorio['PRODUTO'].fillna('')
            df_antigo_relatorio['TIPO'] = df_antigo_relatorio['TIPO'].fillna('')
            df_antigo_relatorio['PROMOÇÃO'] = pd.to_numeric(df_antigo_relatorio['PROMOÇÃO'], errors='coerce')

            if 'ID' in df_antigo_relatorio.columns and df_antigo_relatorio['ID'].notna().any():
                 df_antigo_relatorio['CHAVE_COMP'] = df_antigo_relatorio['ID'].astype(str)
            else:
                 df_antigo_relatorio['CHAVE_COMP'] = df_antigo_relatorio['PRODUTO'].astype(str) + "||" + df_antigo_relatorio['TIPO'].astype(str)

            mapa_precos_antigos = pd.Series(df_antigo_relatorio['PROMOÇÃO'].values, index=df_antigo_relatorio['CHAVE_COMP']).to_dict()
            chaves_antigas_presentes = set(mapa_precos_antigos.keys())

        except FileNotFoundError:
            print(f"ℹ️  Relatório anterior '{FINAL_OFFERS_REPORT_NAME}' não encontrado. Todas as ofertas atuais serão consideradas novas para destaque.")
        except Exception as e:
            print(f"❌ Erro ao ler relatório anterior '{FINAL_OFFERS_REPORT_NAME}': {e}. Criando novo sem destaques de atualização.")

    if mode == 'update' and not df_relatorio_parcial.empty:
        if 'ID' in df_relatorio_parcial.columns and df_relatorio_parcial['ID'].notna().any():
            df_relatorio_parcial['CHAVE_COMP'] = df_relatorio_parcial['ID'].astype(str)
        else:
            df_relatorio_parcial['CHAVE_COMP'] = df_relatorio_parcial['PRODUTO'].astype(str) + "||" + df_relatorio_parcial['TIPO'].astype(str)

        for index, row_atual in df_relatorio_parcial.iterrows():
            chave_atual = row_atual['CHAVE_COMP']
            preco_atual = row_atual['PROMOÇÃO']
            destacar_linha = False
            if chave_atual not in chaves_antigas_presentes:
                destacar_linha = True
            else:
                preco_antigo = mapa_precos_antigos.get(chave_atual)
                if pd.isna(preco_atual) != pd.isna(preco_antigo):
                    destacar_linha = True
                elif not pd.isna(preco_atual) and not pd.isna(preco_antigo) and not np.isclose(preco_atual, preco_antigo, equal_nan=False):
                    destacar_linha = True

            if destacar_linha:
                df_relatorio_parcial.loc[index, 'DESTAQUE'] = True

        df_relatorio_parcial = df_relatorio_parcial.drop(columns=['CHAVE_COMP'], errors='ignore')

    for col in ['NOME_PROMOÇÃO', 'SESSÃO', 'PRODUTO', 'TIPO']:
        if col in df_relatorio_parcial.columns:
            df_relatorio_parcial[col] = df_relatorio_parcial[col].fillna("")
    for col in ['ID', 'PROMOÇÃO']: # Já garantido acima, mas redundância não prejudica
        if col not in df_relatorio_parcial.columns:
            df_relatorio_parcial[col] = np.nan
        else: # Certificar que IDs possam ser NaN onde não existem
            if col == 'ID':
                df_relatorio_parcial[col] = pd.to_numeric(df_relatorio_parcial[col], errors='coerce').astype('Int64')


    # --- INÍCIO DA MODIFICAÇÃO PARA ORDENAÇÃO E ESPAÇAMENTO (COM ESPAÇO ENTRE SEÇÕES TAMBÉM) ---
    if 'SESSÃO' in df_relatorio_parcial.columns:
        df_relatorio_parcial['Chave_Grupo_Secao_Ordenacao'] = df_relatorio_parcial['SESSÃO'].apply(get_section_sort_key)
    else:
        df_relatorio_parcial['Chave_Grupo_Secao_Ordenacao'] = np.nan

    df_ordenado = df_relatorio_parcial.sort_values(
        by=['NOME_PROMOÇÃO', 'Chave_Grupo_Secao_Ordenacao', 'SESSÃO', 'PRODUTO'],
        ascending=[True, True, True, True]
    ).reset_index(drop=True)

    lista_linhas_com_espacos = []
    ultimo_nome_promocao_visto = None
    ultima_chave_grupo_secao_vista = None
    # colunas_df_ordenado_com_destaque agora inclui 'Chave_Grupo_Secao_Ordenacao'
    colunas_df_ordenado_com_destaque = df_ordenado.columns.tolist()

    if not df_ordenado.empty:
        for _, row_data in df_ordenado.iterrows():
            nome_promocao_atual = row_data['NOME_PROMOÇÃO']
            chave_grupo_secao_atual = row_data['Chave_Grupo_Secao_Ordenacao']

            adicionar_espaco = False
            if ultimo_nome_promocao_visto is not None:
                if nome_promocao_atual != ultimo_nome_promocao_visto:
                    adicionar_espaco = True
                elif chave_grupo_secao_atual != ultima_chave_grupo_secao_vista:
                    adicionar_espaco = True

            if adicionar_espaco:
                spacer_row_data = {col: np.nan for col in colunas_df_ordenado_com_destaque}
                if 'DESTAQUE' in spacer_row_data:
                    spacer_row_data['DESTAQUE'] = False
                lista_linhas_com_espacos.append(spacer_row_data)

            lista_linhas_com_espacos.append(row_data.to_dict())

            ultimo_nome_promocao_visto = nome_promocao_atual
            ultima_chave_grupo_secao_vista = chave_grupo_secao_atual
    # --- FIM DA MODIFICAÇÃO PARA ORDENAÇÃO E ESPAÇAMENTO ---

    if not lista_linhas_com_espacos:
        print(f"⚠️ Nenhum dado para gerar o relatório '{FINAL_OFFERS_REPORT_NAME}'.");
        if mode == 'update' and os.path.exists(path_relatorio_final):
            try:
                pd.DataFrame(columns=colunas_finais_excel_ordenadas).to_excel(path_relatorio_final, index=False, sheet_name='Ofertas Finais')
                print(f"ℹ️ Relatório anterior '{FINAL_OFFERS_REPORT_NAME}' substituído por um arquivo vazio (sem ofertas atuais).")
            except Exception as e_save_empty:
                print(f"❌ Erro ao tentar salvar relatório vazio: {e_save_empty}")
        return

    df_output_excel = pd.DataFrame(lista_linhas_com_espacos)
    df_output_excel = df_output_excel.drop(columns=['Chave_Grupo_Secao_Ordenacao'], errors='ignore')


    try:
        with pd.ExcelWriter(path_relatorio_final, engine='openpyxl') as writer:
            # df_para_salvar_excel deve conter apenas as colunas finais para o Excel
            df_para_salvar_excel = df_output_excel[colunas_finais_excel_ordenadas].copy() # Seleciona e copia
            # Certificar que 'DESTAQUE' não está em df_para_salvar_excel (já não deveria estar por colunas_finais_excel_ordenadas)
            # df_para_salvar_excel = df_output_excel.drop(columns=['DESTAQUE'], errors='ignore') # Redundante se usando colunas_finais_excel_ordenadas

            df_para_salvar_excel.to_excel(writer, index=False, sheet_name='Ofertas Finais')
            ws = writer.sheets['Ofertas Finais']

            ws.sheet_view.showGridLines = False
            thin_side = Side(style='thin')
            border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            col_names_in_excel = df_para_salvar_excel.columns.tolist()

            header_row_excel = 1
            default_format_config = COLUMNS_FORMATTING_CONFIG.get('_DEFAULT_', {})
            for col_idx_0based, col_name_excel in enumerate(col_names_in_excel):
                cell = ws.cell(row=header_row_excel, column=col_idx_0based + 1)
                col_config = COLUMNS_FORMATTING_CONFIG.get(col_name_excel, {})

                font_name = col_config.get('header_font_name', default_format_config.get('header_font_name', 'Calibri'))
                font_size = col_config.get('header_font_size', default_format_config.get('header_font_size', 9))
                font_color_hex = col_config.get('header_font_color', default_format_config.get('header_font_color', "000000")).replace("#","")
                font_bold = col_config.get('header_font_bold', default_format_config.get('header_font_bold', True))
                font_italic = col_config.get('header_font_italic', default_format_config.get('header_font_italic', False))
                fill_color_hex = col_config.get('header_fill_color', default_format_config.get('header_fill_color', None))
                if fill_color_hex: fill_color_hex = fill_color_hex.replace("#","")
                align_h = col_config.get('header_alignment_horizontal', default_format_config.get('header_alignment_horizontal', 'left'))
                align_v = col_config.get('header_alignment_vertical', default_format_config.get('header_alignment_vertical', 'center'))

                cell.font = Font(name=font_name, size=font_size, color=font_color_hex, bold=font_bold, italic=font_italic)
                if fill_color_hex:
                    cell.fill = PatternFill(start_color=fill_color_hex, end_color=fill_color_hex, fill_type="solid")
                cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=False)
                cell.border = border_style

            # Acessar a coluna DESTAQUE de df_output_excel (que a tem) para formatar as linhas de dados
            for df_row_idx, (_, df_row_data_com_destaque) in enumerate(df_output_excel.iterrows()):
                row_idx_excel = df_row_idx + 2 # +1 para cabeçalho, +1 para 0-based to 1-based
                # Verificar se é linha de espaçamento usando df_row_data_com_destaque
                is_spacer_row_df = all(pd.isna(val) for col, val in df_row_data_com_destaque.items() if col not in ['DESTAQUE', 'Chave_Grupo_Secao_Ordenacao'])
                if is_spacer_row_df: continue

                deve_destacar_linha = df_row_data_com_destaque.get('DESTAQUE', False) == True

                for col_idx_0based, col_name_excel in enumerate(col_names_in_excel): # col_names_in_excel são as colunas do df_para_salvar_excel
                    cell = ws.cell(row=row_idx_excel, column=col_idx_0based + 1)
                    col_config = COLUMNS_FORMATTING_CONFIG.get(col_name_excel, {})

                    base_font_name = col_config.get('font_name', default_format_config.get('font_name', 'Calibri'))
                    base_font_size = col_config.get('font_size', default_format_config.get('font_size', 10))
                    base_font_color_hex = col_config.get('font_color', default_format_config.get('font_color', "000000")).replace("#","")
                    base_font_bold = col_config.get('font_bold', default_format_config.get('font_bold', False))
                    base_font_italic = col_config.get('font_italic', default_format_config.get('font_italic', False))
                    base_fill_color_hex = col_config.get('fill_color', default_format_config.get('fill_color', None))
                    if base_fill_color_hex: base_fill_color_hex = base_fill_color_hex.replace("#","")

                    number_fmt = col_config.get('number_format', default_format_config.get('number_format', '@'))
                    align_h = col_config.get('alignment_horizontal', default_format_config.get('alignment_horizontal', 'left'))
                    align_v = col_config.get('alignment_vertical', default_format_config.get('alignment_vertical', 'center'))

                    final_font = Font(name=base_font_name, size=base_font_size, color=base_font_color_hex, bold=base_font_bold, italic=base_font_italic)
                    final_fill = PatternFill(fill_type=None)
                    if base_fill_color_hex:
                        final_fill = PatternFill(start_color=base_fill_color_hex, end_color=base_fill_color_hex, fill_type="solid")

                    if deve_destacar_linha:
                        final_fill = HIGHLIGHT_FILL
                        final_font = Font(
                            name=HIGHLIGHT_FONT.name if HIGHLIGHT_FONT.name else base_font_name,
                            size=HIGHLIGHT_FONT.size if HIGHLIGHT_FONT.size else base_font_size,
                            color=HIGHLIGHT_FONT.color.rgb if HIGHLIGHT_FONT.color else base_font_color_hex,
                            bold=HIGHLIGHT_FONT.bold if HIGHLIGHT_FONT.bold is not None else base_font_bold,
                            italic=HIGHLIGHT_FONT.italic if HIGHLIGHT_FONT.italic is not None else base_font_italic
                        )

                    cell.font = final_font
                    cell.fill = final_fill
                    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=False)
                    cell.border = border_style

                    if cell.value is not None:
                        if number_fmt not in ['@', 'General'] and isinstance(cell.value, str):
                            try: cell.value = float(str(cell.value).replace("R$", "").replace(".", "").replace(",", ".").strip())
                            except ValueError: pass

                        if number_fmt == 'R$ #,##0.00' and isinstance(cell.value, (int, float)) and cell.value == 0:
                            cell.value = 0.00
                        cell.number_format = number_fmt

            for col_idx_openpyxl_0based, col_name_excel in enumerate(col_names_in_excel):
                column_letter = get_column_letter(col_idx_openpyxl_0based + 1)
                max_len = 0

                current_col_config = COLUMNS_FORMATTING_CONFIG.get(col_name_excel, {})
                current_number_format = current_col_config.get('number_format', default_format_config.get('number_format', '@'))

                for cell_row_idx_ws in range(1, ws.max_row + 1):
                    cell = ws.cell(row=cell_row_idx_ws, column=col_idx_openpyxl_0based + 1)
                    if cell.value is not None:
                        text_val = str(cell.value)
                        if isinstance(cell.value, (int, float)) and current_number_format == 'R$ #,##0.00':
                            text_val = f"R$ {cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        elif isinstance(cell.value, (int,float)) and current_number_format == '0':
                            text_val = f"{cell.value:.0f}"

                        current_len = len(text_val)
                        max_len = max(max_len, current_len)

                header_cell_val = ws.cell(row=1, column=col_idx_openpyxl_0based + 1).value
                if header_cell_val:
                    max_len = max(max_len, len(str(header_cell_val)))

                adjusted_width = min(max_len + 4, 60)
                ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width > 8 else 10

        num_ofertas_reais = len(df_ordenado.dropna(subset=['PRODUTO'], how='all')) # df_ordenado ainda tem a forma antes dos espaçadores
        print(f"✅ Relatório '{FINAL_OFFERS_REPORT_NAME}' criado com {num_ofertas_reais} ofertas na raiz do projeto.")
        if mode == 'update':
            # df_ordenado tem 'DESTAQUE' e 'Chave_Grupo_Secao_Ordenacao'
            num_destacados = df_ordenado['DESTAQUE'].sum() if 'DESTAQUE' in df_ordenado.columns else 0
            if df_antigo_relatorio is not None :
                 print(f"   ✨ {num_destacados} ofertas foram destacadas como novas ou com preço alterado.")
            elif not chaves_antigas_presentes and num_ofertas_reais > 0:
                 print(f"   ✨ Todas as {num_ofertas_reais} ofertas são consideradas novas (nenhum relatório anterior para comparar).")

    except Exception as e:
        print(f"❌ Erro ao salvar/formatar Excel '{FINAL_OFFERS_REPORT_NAME}': {e}")
        traceback.print_exc()

# ------------------------------------------
# FUNÇÃO PRINCIPAL
# ------------------------------------------

def main_logic(mode: str = 'create') -> None:
    print(f"🚀 Iniciando processamento de ofertas (Modo: {mode})...")

    if not os.path.isdir(INPUT_XML_DIR):
        print(f"❌ ERRO: Diretório de entrada '{INPUT_XML_DIR}' não encontrado. Crie-o e adicione os arquivos XML.")
        print("🏁 Processamento abortado.")
        return

    arquivos_xml_originais = listar_arquivos_xml(INPUT_XML_DIR)
    if not arquivos_xml_originais:
        print(f"ℹ️ Nenhum arquivo XML encontrado em '{INPUT_XML_DIR}'.")
    else:
        print(f"🔄 Normalizando nomes de {len(arquivos_xml_originais)} arquivos XML...")
        renomear_arquivos_em_lote(arquivos_xml_originais, INPUT_XML_DIR)

    arquivos_xml_para_processar = [
        f for f in os.listdir(INPUT_XML_DIR)
        if f.endswith(XML_EXTENSION) and os.path.isfile(os.path.join(INPUT_XML_DIR, f))
    ]
    if not arquivos_xml_para_processar:
        print(f"ℹ️ Nenhum arquivo XML para processar em '{INPUT_XML_DIR}' após normalização.")

    print(f"📄 Consolidando dados de {len(arquivos_xml_para_processar)} arquivos XML...")
    df_ofertas_consolidadas_xml = consolidar_dados_de_xmls(arquivos_xml_para_processar)

    if df_ofertas_consolidadas_xml.empty and not arquivos_xml_para_processar:
        print("ℹ️ Nenhum dado de produto foi consolidado dos arquivos XML nesta execução.")

    print(f"💾 Atualizando o arquivo de cadastro de produtos: './{PRODUCT_MASTER_FILE_NAME}'...")
    atualizar_cadastro_produtos_base(df_ofertas_consolidadas_xml)

    gerar_relatorio = True
    path_relatorio_final_check = f'./{FINAL_OFFERS_REPORT_NAME}' # Usar uma variável local para o check
    if df_ofertas_consolidadas_xml.empty:
        if mode == 'create':
            print("🏁 Processamento concluído (sem dados XML para gerar relatório de ofertas no modo 'create').")
            gerar_relatorio = False
        elif mode == 'update' and not os.path.exists(path_relatorio_final_check):
            print("🏁 Processamento concluído (sem dados XML e sem relatório anterior para atualizar).")
            gerar_relatorio = False

    if gerar_relatorio:
        print(f"📊 Gerando o relatório final de ofertas: './{FINAL_OFFERS_REPORT_NAME}'...")
        gerar_relatorio_ofertas_finais(df_ofertas_consolidadas_xml, mode)

    print("🏁 Processamento concluído.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Processa arquivos XML de ofertas e gera relatórios Excel.")
    parser.add_argument(
        "--mode",
        choices=['create', 'update'],
        default='create',
        help="Modo de operação: 'create' para gerar do zero (padrão), 'update' para destacar mudanças no relatório de ofertas."
    )
    args = parser.parse_args()
    main_logic(mode=args.mode)