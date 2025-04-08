import openpyxl
import requests
import json
import os
import time
import random
from dotenv import load_dotenv

# Configuração
load_dotenv()
API_KEY = os.getenv('GEMINI_API_KEY')
BASE_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={API_KEY}"

# Cache para evitar repetições
product_cache = {}

class RateLimiter:
    def __init__(self):
        self.last_call = 0
        self.delay = 7  # Aumentado para evitar rate limit
        self.retry_delay = 30

    def wait(self):
        """Controla o rate limiting com jitter"""
        elapsed = time.time() - self.last_call
        if elapsed < self.delay:
            wait_time = self.delay - elapsed + random.uniform(0, 2)
            time.sleep(wait_time)
        self.last_call = time.time()

rate_limiter = RateLimiter()

def call_gemini_api(prompt):
    """Chamada à API com tratamento robusto de erros"""
    headers = {'Content-Type': 'application/json'}
    data = {
        "contents": [{
            "parts": [{"text": prompt}]
        }],
        "generationConfig": {
            "temperature": 0.1,  # Mais determinístico
            "topP": 0.9,
            "topK": 20,
            "maxOutputTokens": 300
        }
    }

    for attempt in range(3):
        rate_limiter.wait()
        try:
            response = requests.post(BASE_URL, headers=headers, json=data, timeout=30)

            if response.status_code == 429:
                wait_time = self.retry_delay * (attempt + 1)
                print(f"Rate limit excedido. Aguardando {wait_time} segundos...")
                time.sleep(wait_time)
                continue

            response.raise_for_status()
            result = response.json()
            return result['candidates'][0]['content']['parts'][0]['text'].strip()

        except Exception as e:
            print(f"Erro na chamada API (tentativa {attempt + 1}): {str(e)[:200]}")
            if attempt < 2:
                time.sleep(10 * (attempt + 1))

    return None

def format_product_name(product_name):
    """Reformata o nome do produto sem hífens e com ordem específica"""
    cache_key = f"product_{product_name}"
    if cache_key in product_cache:
        return product_cache[cache_key]

    prompt = f"""Reformate este nome de produto seguindo ESTRITAMENTE:
    [TIPO] [MARCA] [MEDIDA] [ESPECIFICAÇÃO]
    (SEM hífens, separado por espaços simples)

    Exemplos CORRETOS:
    - 'ACHOCOLATADO EM PÓ 3 CORACOES CHOCOLATTO SACHÊ 400G' → 'ACHOCOLATADO EM PÓ 3 CORACOES SACHÊ 400G CHOCOLATTO'
    - 'ARROZ BRILHANTE BRANCO TIPO 1 PCT 5KG' → 'ARROZ BRILHANTE PCT 5KG BRANCO TIPO 1'
    - 'DESODORANTE MOOD AEROSOL FEMININO HIDRATANTE PROTECT 150ML' → 'DESODORANTE MOOD AEROSOL 150ML FEMININO HIDRATANTE PROTECT'

    Regras CRÍTICAS:
    1. NUNCA use hífens
    2. ORDEM: TIPO MARCA MEDIDA ESPECIFICAÇÃO
    3. MEDIDA deve incluir a unidade (G, KG, ML, L, etc)
    4. Mantenha TODAS informações originais

    Produto para reformatar: '{product_name}'

    Responda APENAS com o nome formatado, SEM comentários."""

    result = call_gemini_api(prompt)

    # Validação e limpeza
    if result:
        # Remove hífens residuais
        result = result.replace(" - ", " ").replace("-", " ")
        # Remove múltiplos espaços
        result = " ".join(result.split())

        # Verifica se manteve as informações essenciais
        original_words = product_name.split()
        if all(word in result for word in original_words[:2]):
            product_cache[cache_key] = result
            return result

    return product_name  # Retorna original se falhar

def classify_section(current_section, product_name):
    """Mantém a numeração existente da seção"""
    # Extrai a numeração existente (ex: #01 MERCEARIA - #01 ALTO GIRO → mantém igual)
    return current_section

def process_worksheet(ws):
    """Processamento da planilha com checkpoint"""
    headers = [cell.value for cell in ws[1]]

    try:
        section_idx = headers.index("Seção")
        product_idx = headers.index("Produto")
        corrected_idx = headers.index("Produto Corrigido")
    except ValueError as e:
        print(f"Erro nas colunas: {e}")
        return

    # Sistema de checkpoint
    start_row = 2
    if os.path.exists('checkpoint.json'):
        with open('checkpoint.json') as f:
            checkpoint = json.load(f)
            start_row = checkpoint.get('row', 2)
            product_cache.update(checkpoint.get('cache', {}))

    for row in ws.iter_rows(min_row=start_row, values_only=False):
        try:
            # NÃO altera a seção (mantém numeração original)

            # Reformata nome do produto
            if row[corrected_idx].value:
                original_name = row[corrected_idx].value
                formatted_name = format_product_name(original_name)

                # Validação rigorosa
                if (formatted_name and
                    formatted_name != original_name and
                    len(formatted_name.split()) >= 4 and  # Mínimo 4 componentes
                    not "-" in formatted_name):
                    row[corrected_idx].value = formatted_name
                else:
                    print(f"Formatação mantida original para: {original_name}")

            # Checkpoint a cada 5 linhas
            if row[0].row % 5 == 0:
                with open('checkpoint.json', 'w') as f:
                    json.dump({
                        'row': row[0].row,
                        'cache': product_cache
                    }, f)

        except Exception as e:
            print(f"Erro na linha {row[0].row}: {str(e)[:200]}")
            continue

def main():
    """Função principal"""
    try:
        workbook = openpyxl.load_workbook("dados_produtos_corrigidos.xlsx")

        for sheet_name in workbook.sheetnames:
            print(f"\nProcessando: {sheet_name}")
            process_worksheet(workbook[sheet_name])

        workbook.save("dados_produtos_corrigidos.xlsx")

        # Limpeza final
        if os.path.exists('checkpoint.json'):
            os.remove('checkpoint.json')

        print("\nProcessamento concluído com sucesso!")
        print(f"Total de itens processados: {len(product_cache)}")

    except KeyboardInterrupt:
        print("\nProcessamento interrompido. Pode retomar depois.")
    except Exception as e:
        print(f"\nErro fatal: {str(e)[:200]}")
    finally:
        if 'workbook' in locals():
            workbook.close()

if __name__ == "__main__":
    main()