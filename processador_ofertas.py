# -*- coding: utf-8 -*-
from typing import List, Dict, Any
import pandas as pd
import os
import glob
import xmltodict
from typing import List, Dict, Any, Optional, Tuple, Set
import re
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import traceback
import argparse

# ------------------------------------------
# CONSTANTES E CONFIGURAÇÕES GLOBAIS
# ------------------------------------------
DIRETORIO_ENTRADA_XML = './arquivos_xml_entrada/'
EXTENSAO_XML = '.xml'
NOME_ARQUIVO_MESTRE_PRODUTOS = 'produtos_cadastrados.xlsx'
NOME_RELATORIO_FINAL_OFERTAS = 'giro_da_praça_ofertas.xlsx'

SECOES_EXCECAO_AGRUPAMENTO = ["#01 MERCEARIA - #01 ALTO GIRO", "#01 MERCEARIA - #02 ALTO GIRO", "#02 MERCEARIA - LATICÍNIOS - LEITE"]

PREENCHIMENTO_DESTAQUE = PatternFill(start_color="45A045", end_color="45A045", fill_type="solid")
FONTE_DESTAQUE = Font(color="FFFFFF", bold=True)

# --- NOVAS CONSTANTES PARA ORDENAÇÃO DO RELATÓRIO FINAL ---
# Coluna principal para ordenação.
# Opções válidas (nomes das colunas no DataFrame final antes de ir para o Excel):
# 'NOME_PROMOÇÃO', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO'
COLUNA_ORDENACAO_PRIMARIA_RELATORIO = 'NOME_PROMOÇÃO'
# Define a ordem da coluna primária: True para ascendente, False para descendente.
ORDEM_ASCENDENTE_PRIMARIA_RELATORIO = True
# -----------------------------------------------------------

CONFIGURACAO_FORMATACAO_COLUNAS = {
    '_DEFAULT_': {
        'font_name': 'Calibri',
        'font_size': 10,
        'font_color': "000000",
        'font_bold': False,
        'font_italic': False,
        'number_format': '@',
        'alignment_horizontal': 'left',
        'alignment_vertical': 'center',
        'fill_color': None,
        'header_font_name': 'Calibri',
        'header_font_size': 9,
        'header_font_color': "000000",
        'header_font_bold': True,
        'header_font_italic': False,
        'header_fill_color': "FFFFFF",
        'header_alignment_horizontal': 'left',
        'header_alignment_vertical': 'center',
    },
    'NOME_PROMOÇÃO': {'font_size': 7},
    'SESSÃO': {'font_size': 7, 'font_color': "000000"},
    'ID': {'font_size': 7, 'number_format': '0'},
    'PRODUTO': {'font_size': 9},
    'TIPO': {'font_size': 9, 'font_color': "000000", 'font_italic': True},
    'PROMOÇÃO': {
        'font_size': 10,
        'font_color': "000000",
        'font_bold': True,
        'number_format': 'R$ #,##0.00',
        'alignment_horizontal': 'right',
    }
}

# ------------------------------------------
# FUNÇÕES UTILITÁRIAS DE TEXTO (PURAS)
# ------------------------------------------


def remover_acentos_de_texto(texto: str) -> str:
    if not isinstance(texto, str):
        return ""
    mapeamento_acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u', 'ã': 'a', 'õ': 'o', 'ä': 'a', 'ë': 'e', 'ï': 'i',
        'ö': 'o', 'ü': 'u', 'ç': 'c', 'ñ': 'n', 'ÿ': 'y', 'ý': 'y', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O',
        'Ú': 'U', 'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U', 'Â': 'A', 'Ê': 'E', 'Î': 'I', 'Ô': 'O',
        'Û': 'U', 'Ã': 'A', 'Õ': 'O', 'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U', 'Ç': 'C', 'Ñ': 'N',
        'Ÿ': 'Y', 'Ý': 'Y'
    }
    return ''.join(mapeamento_acentos.get(c, c) for c in texto)


def normalizar_texto_para_comparacao(texto: str) -> str:
    if not isinstance(texto, str):
        return ""
    texto_sem_acentos = remover_acentos_de_texto(texto)
    return texto_sem_acentos.lower()

# ------------------------------------------
# FUNÇÕES UTILITÁRIAS DE SISTEMA DE ARQUIVOS (COM EFEITOS COLATERAIS)
# ------------------------------------------


def listar_arquivos_por_extensao(diretorio: str, extensao: str) -> List[str]:
    return glob.glob(os.path.join(diretorio, f'*{extensao}'))


def renomear_arquivo_individual(caminho_antigo: str, novo_diretorio: str) -> None:
    nome_arquivo_base = os.path.basename(caminho_antigo)
    novo_nome_base = normalizar_texto_para_comparacao(nome_arquivo_base)
    novo_caminho_completo = os.path.join(novo_diretorio, novo_nome_base)
    try:
        if caminho_antigo != novo_caminho_completo:
            os.rename(caminho_antigo, novo_caminho_completo)
    except FileExistsError:
        print(f"⚠️  Arquivo '{novo_caminho_completo}' já existe. Não foi renomeado '{caminho_antigo}'.")
    except Exception as e:
        print(f"❌ Erro ao renomear '{caminho_antigo}' para '{novo_caminho_completo}': {e}")


def normalizar_nomes_arquivos_xml_em_lote(arquivos_xml_entrada: List[str], diretorio_xml: str) -> None:
    print(f"🔄 Normalizando nomes de {len(arquivos_xml_entrada)} arquivos XML...")
    for caminho_arquivo in arquivos_xml_entrada:
        renomear_arquivo_individual(caminho_arquivo, diretorio_xml)

# ------------------------------------------
# FUNÇÕES DE PROCESSAMENTO DE XML
# ------------------------------------------


def converter_dados_produto_xml_para_dicionario(dados_produto_xml: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "Promoção_XML": dados_produto_xml.get('descrpromocao', ''),
        "ID_Produto_XML": int(dados_produto_xml.get('idsubproduto', 0)),
        "Nome_Produto_XML": dados_produto_xml.get('descrresproduto', ''),
        "Preco_Promocao_XML": float(dados_produto_xml.get('precopromocao', 0.0) if dados_produto_xml.get('precopromocao') not in [None, ''] else 0.0)
    }


def ler_e_extrair_produtos_de_arquivo_xml(
    nome_arquivo_xml: str,
    diretorio_xml: str
) -> List[Dict[str, Any]]:

    caminho_completo_arquivo = os.path.join(diretorio_xml, nome_arquivo_xml)

    try:
        with open(caminho_completo_arquivo, 'rb') as arquivo_xml:
            dados_xml_lidos = xmltodict.parse(arquivo_xml)

        # Procura dinamicamente pela chave que começa com "temporario_846"
        chave_principal = next(
            (
                chave
                for chave in dados_xml_lidos.keys()
                if chave.startswith("temporario_846")
            ),
            None
        )

        if not chave_principal:
            print(
                f"⚠️ Nenhuma estrutura iniciando com "
                f"'temporario_846' encontrada em {nome_arquivo_xml}."
            )
            return []

        conteudo_principal = dados_xml_lidos[chave_principal]

        # Procura dinamicamente pela chave de linhas (_row)
        chave_rows = next(
            (
                chave
                for chave in conteudo_principal.keys()
                if chave.startswith(chave_principal) and chave.endswith("_row")
            ),
            None
        )

        if not chave_rows:
            print(
                f"⚠️ Nenhuma estrutura de linhas encontrada "
                f"em {nome_arquivo_xml}."
            )
            return []

        lista_produtos_xml_bruto = conteudo_principal.get(chave_rows, [])

        # Quando existe apenas um item, xmltodict retorna dict ao invés de lista
        if isinstance(lista_produtos_xml_bruto, dict):
            lista_produtos_xml_bruto = [lista_produtos_xml_bruto]

        if not lista_produtos_xml_bruto:
            return []

        return [
            converter_dados_produto_xml_para_dicionario(produto)
            for produto in lista_produtos_xml_bruto
            if isinstance(produto, dict)
        ]

    except FileNotFoundError:
        print(f"❌ Arquivo XML não encontrado: {caminho_completo_arquivo}")
        return []

    except Exception as e:
        print(f"❌ Erro ao processar XML {nome_arquivo_xml}: {e}")
        traceback.print_exc()
        return []


def consolidar_dados_de_multiplos_arquivos_xml(lista_nomes_arquivos_xml: List[str], diretorio_xml: str) -> pd.DataFrame:
    lista_global_produtos = []
    print(f"📄 Consolidando dados de {len(lista_nomes_arquivos_xml)} arquivos XML...")
    for nome_arquivo in lista_nomes_arquivos_xml:
        produtos_deste_arquivo = ler_e_extrair_produtos_de_arquivo_xml(nome_arquivo, diretorio_xml)
        if produtos_deste_arquivo:
            lista_global_produtos.extend(produtos_deste_arquivo)
            print(f"✅ {nome_arquivo}: {len(produtos_deste_arquivo)} produtos processados.")

    if not lista_global_produtos:
        print("⚠️ Nenhum produto encontrado nos arquivos XML.")
        return pd.DataFrame()
    return pd.DataFrame(lista_global_produtos)

# ------------------------------------------
# FUNÇÕES AUXILIARES DE TRATAMENTO DE DADOS DE PRODUTO (PURAS)
# ------------------------------------------


def extrair_nome_base_produto(nome_completo_produto: str) -> str:
    if not isinstance(nome_completo_produto, str):
        return ""
    unidades = [
        "KILOGRAMAS", "KILOGRAMA", "QUILOGRAMAS", "QUILOGRAMA", "LITROS", "LITRO", "UNIDADES", "UNIDADE",
        "PACOTES", "PACOTE", "CAIXAS", "CAIXA", "FARDOS", "FARDO", "GARRAFAS", "GARRAFA", "LATAS", "LATA",
        "POTES", "POTE", "ROLOS", "ROLO", "SACHES", "SACHE", "TUBOS", "TUBO", "VIDROS", "VIDRO", "SACOLAS",
        "SACOLA", "BLISTERS", "BLISTER", "CARTELAS", "CARTELA", "DISPLAYS", "DISPLAY", "GALÕES", "GALÃO",
        "REFILs", "REFIL", "TABLETES", "TABLETE", "PEÇAS", "PEÇA", "FOLHAS", "FOLHA", "FLACONETES",
        "FLACONETE", "BISNAGAS", "BISNAGA", "AMPOLAS", "AMPOLA", "DRAGEAS", "CAPSULAS", "CAPSULA",
        "COMPRIMIDOS", "COMPRIMIDO", "ENVELOPES", "ENVELOPE",
        "KG", "GR", "G", "MG", "L", "ML", "CM", "MM", "M",
        "UN", "UND", "UNID", "PC", "PÇ", "PCT", "CX", "FD", "GF", "LT", "PT", "RL", "SC", "TB", "VD",
        "BL", "CT", "DP", "GL", "RF", "TBL", "FL", "CAPS", "COMP", "ENV", "KIT",
    ]
    nome_produto_normalizado_espacos = ' '.join(nome_completo_produto.split())
    padrao_regex_unidades = "|".join(map(re.escape, sorted(unidades, key=len, reverse=True)))
    expressao_regular = r"^(.*?)(\s?)(\d+[\.,]?\d*)(\s*)(" + padrao_regex_unidades + r")($|\s+.*$)"
    match = re.match(expressao_regular, nome_produto_normalizado_espacos, re.IGNORECASE)

    if match:
        parte_antes_medida = match.group(1).strip()
        quantidade_texto = match.group(3)
        unidade_texto = match.group(5)
        nome_base = f"{parte_antes_medida} {quantidade_texto}{unidade_texto}".strip()
        return ' '.join(nome_base.split())
    return nome_produto_normalizado_espacos


def extrair_tipo_variacao_produto(nome_completo_produto: str, nome_base_produto: str) -> str:
    if not isinstance(nome_completo_produto, str) or not isinstance(nome_base_produto, str):
        return ""
    nome_completo_norm = ' '.join(nome_completo_produto.split())
    nome_base_norm = ' '.join(nome_base_produto.split())

    if nome_completo_norm == nome_base_norm:
        return ""
    if nome_completo_norm.startswith(nome_base_norm):
        parte_tipo = nome_completo_norm[len(nome_base_norm):].strip()
        parte_tipo = re.sub(r"^[ \-\/\_]+", "", parte_tipo).strip()
        return parte_tipo
    return ""


def obter_chave_ordenacao_secao(nome_secao: Any) -> str:
    if not isinstance(nome_secao, str) or pd.isna(nome_secao):
        return "SEM SEÇÃO DEFINIDA"
    if nome_secao.startswith("#01 MERCEARIA - #") and "ALTO GIRO" in nome_secao:
        return "#01 MERCEARIA - ALTO GIRO"
    return nome_secao.split(" - ", 1)[0]

# ------------------------------------------
# MÓDULO DE GERENCIAMENTO DO CADASTRO DE PRODUTOS
# ------------------------------------------


def carregar_ou_inicializar_dataframe_mestre(caminho_arquivo_mestre: str, colunas_mestre_padrao: List[str]) -> pd.DataFrame:
    try:
        df_mestre = pd.read_excel(caminho_arquivo_mestre)
        if 'ID' not in df_mestre.columns:
            print(f"⚠️  Arquivo '{NOME_ARQUIVO_MESTRE_PRODUTOS}' não possui coluna 'ID'. Tratando como novo.")
            return pd.DataFrame(columns=colunas_mestre_padrao)

        df_temp_mestre = pd.DataFrame(columns=colunas_mestre_padrao)
        for coluna in colunas_mestre_padrao:
            if coluna in df_mestre.columns:
                df_temp_mestre[coluna] = df_mestre[coluna]
            else:
                df_temp_mestre[coluna] = np.nan
        return df_temp_mestre
    except FileNotFoundError:
        print(f"ℹ️  Arquivo '{NOME_ARQUIVO_MESTRE_PRODUTOS}' não encontrado. Será criado.")
        return pd.DataFrame(columns=colunas_mestre_padrao)
    except Exception as e:
        print(f"❌ Erro ao carregar '{NOME_ARQUIVO_MESTRE_PRODUTOS}': {e}. Operação cancelada para este arquivo.")
        return pd.DataFrame(columns=colunas_mestre_padrao)


def identificar_produtos_novos_para_cadastro(df_ofertas: pd.DataFrame, df_mestre_existente: pd.DataFrame) -> pd.DataFrame:
    if df_ofertas.empty:
        return pd.DataFrame()
    if 'ID_Produto_XML' not in df_ofertas.columns or 'Nome_Produto_XML' not in df_ofertas.columns:
        print(f"❌ Colunas 'ID_Produto_XML' ou 'Nome_Produto_XML' ausentes no DataFrame de ofertas. Não é possível identificar novos produtos.")
        return pd.DataFrame()

    ids_existentes_no_mestre = set()
    if 'ID' in df_mestre_existente.columns and not df_mestre_existente.empty:
        df_mestre_existente['ID'] = pd.to_numeric(df_mestre_existente['ID'], errors='coerce')
        ids_existentes_no_mestre = set(df_mestre_existente['ID'].dropna().astype(int).unique())

    df_ofertas_copia = df_ofertas.copy()
    df_ofertas_copia['ID_Produto_XML'] = pd.to_numeric(df_ofertas_copia['ID_Produto_XML'], errors='coerce')

    df_produtos_novos = df_ofertas_copia[
        ~df_ofertas_copia['ID_Produto_XML'].dropna().astype(int).isin(ids_existentes_no_mestre)
    ].copy()
    return df_produtos_novos


def preparar_novos_produtos_para_dataframe_mestre(df_novos_identificados: pd.DataFrame, colunas_mestre_padrao: List[str]) -> pd.DataFrame:
    if df_novos_identificados.empty:
        return pd.DataFrame(columns=colunas_mestre_padrao)

    lista_novos_formatados = []
    for _, linha in df_novos_identificados.iterrows():
        lista_novos_formatados.append({
            'ID': linha['ID_Produto_XML'],
            'NOME_SISTEMA': linha['Nome_Produto_XML'],
            'SESSÃO': 'SEÇÃO NÃO ESPECIFICADA',
            'NOME_CORRIGIDO': linha['Nome_Produto_XML']
        })
    return pd.DataFrame(lista_novos_formatados, columns=colunas_mestre_padrao)


def salvar_dataframe_mestre_formatado(df_mestre_final: pd.DataFrame, caminho_arquivo_mestre: str) -> None:
    try:
        if 'ID' in df_mestre_final.columns:
            df_mestre_final['ID'] = pd.to_numeric(df_mestre_final['ID'], errors='coerce')
        df_mestre_final.to_excel(caminho_arquivo_mestre, index=False)
    except Exception as e:
        raise Exception(f"Erro ao salvar DataFrame mestre em '{caminho_arquivo_mestre}': {e}")


def aplicar_estilos_visuais_arquivo_mestre(caminho_arquivo_mestre: str, ids_produtos_em_oferta: Set[int], preenchimento_destaque: PatternFill) -> List[Dict[str, Any]]:
    produtos_coloridos_info = []
    try:
        workbook = load_workbook(caminho_arquivo_mestre)
        planilha = workbook.active
        sem_preenchimento = PatternFill(fill_type=None)

        id_col_idx, nome_corrigido_col_idx, nome_sistema_col_idx = None, None, None
        if planilha.max_row > 0:
            celulas_cabecalho = planilha[1]
            for i, celula in enumerate(celulas_cabecalho):
                if celula.value == 'ID':
                    id_col_idx = i + 1
                elif celula.value == 'NOME_CORRIGIDO':
                    nome_corrigido_col_idx = i + 1
                elif celula.value == 'NOME_SISTEMA':
                    nome_sistema_col_idx = i + 1

        if id_col_idx is None and planilha.max_row > 0:
            print(f"⚠️  Aviso: Coluna 'ID' não encontrada no cabeçalho de '{NOME_ARQUIVO_MESTRE_PRODUTOS}'. Não será possível aplicar cores.")
            workbook.save(caminho_arquivo_mestre)
            return []

        if planilha.max_row > 1:
            for num_linha in range(2, planilha.max_row + 1):
                for num_coluna in range(1, planilha.max_column + 1):
                    planilha.cell(row=num_linha, column=num_coluna).fill = sem_preenchimento

                valor_celula_id = planilha.cell(row=num_linha, column=id_col_idx).value
                if valor_celula_id is not None:
                    try:
                        id_atual = int(float(valor_celula_id))
                        if id_atual in ids_produtos_em_oferta:
                            for num_coluna in range(1, planilha.max_column + 1):
                                planilha.cell(row=num_linha, column=num_coluna).fill = preenchimento_destaque

                            nome_produto = "NOME NÃO ENCONTRADO"
                            if nome_corrigido_col_idx and planilha.cell(row=num_linha, column=nome_corrigido_col_idx).value:
                                nome_produto = planilha.cell(row=num_linha, column=nome_corrigido_col_idx).value
                            elif nome_sistema_col_idx and planilha.cell(row=num_linha, column=nome_sistema_col_idx).value:
                                nome_produto = planilha.cell(row=num_linha, column=nome_sistema_col_idx).value
                            produtos_coloridos_info.append({'ID': id_atual, 'Nome': nome_produto})
                    except (ValueError, TypeError):
                        pass

        workbook.save(caminho_arquivo_mestre)
        return produtos_coloridos_info
    except Exception as e:
        print(f"❌ Erro ao aplicar estilos visuais em '{caminho_arquivo_mestre}': {e}")
        traceback.print_exc()
        return []


def gerenciar_atualizacao_arquivo_mestre_produtos(df_ofertas_consolidadas_xml: pd.DataFrame) -> None:
    print(f"💾 Atualizando o arquivo de cadastro de produtos: './{NOME_ARQUIVO_MESTRE_PRODUTOS}'...")
    caminho_mestre = f'./{NOME_ARQUIVO_MESTRE_PRODUTOS}'
    colunas_mestre_ordenadas = ['ID', 'SESSÃO', 'NOME_SISTEMA', 'NOME_CORRIGIDO']

    df_mestre_atual = carregar_ou_inicializar_dataframe_mestre(caminho_mestre, colunas_mestre_ordenadas)

    df_novos_produtos_identificados = identificar_produtos_novos_para_cadastro(df_ofertas_consolidadas_xml, df_mestre_atual)

    if not df_novos_produtos_identificados.empty:
        print(f"ℹ️ {len(df_novos_produtos_identificados)} novos produtos encontrados para adicionar.")
        df_novos_formatados = preparar_novos_produtos_para_dataframe_mestre(df_novos_produtos_identificados, colunas_mestre_ordenadas)
        df_mestre_atual = pd.concat([df_mestre_atual, df_novos_formatados], ignore_index=True)
        df_mestre_atual.drop_duplicates(subset=['ID'], keep='first', inplace=True)
    else:
        if not df_ofertas_consolidadas_xml.empty:
            print(f"ℹ️ Nenhum produto novo das ofertas XML para adicionar ao '{NOME_ARQUIVO_MESTRE_PRODUTOS}'.")

    df_mestre_para_salvar = df_mestre_atual.reindex(columns=colunas_mestre_ordenadas)
    df_mestre_para_salvar = df_mestre_para_salvar.sort_values(by=['SESSÃO', 'NOME_SISTEMA', 'ID'])

    try:
        salvar_dataframe_mestre_formatado(df_mestre_para_salvar, caminho_mestre)

        ids_produtos_oferta_set = set()
        if not df_ofertas_consolidadas_xml.empty and 'ID_Produto_XML' in df_ofertas_consolidadas_xml.columns:
            ids_produtos_oferta_set = set(df_ofertas_consolidadas_xml['ID_Produto_XML'].dropna().astype(int).unique())

        if not df_mestre_para_salvar.empty or os.path.exists(caminho_mestre):
            produtos_destacados_info = aplicar_estilos_visuais_arquivo_mestre(caminho_mestre, ids_produtos_oferta_set, PREENCHIMENTO_DESTAQUE)

            if not df_novos_produtos_identificados.empty:
                print(f"✅ '{NOME_ARQUIVO_MESTRE_PRODUTOS}' atualizado com {len(df_novos_produtos_identificados)} novos produtos e formatação de cores aplicada.")
            else:
                print(f"✅ '{NOME_ARQUIVO_MESTRE_PRODUTOS}' salvo com formatação de cores aplicada.")

            if produtos_destacados_info:
                print(f"ℹ️ Produtos das ofertas XML (coloridos em verde em '{NOME_ARQUIVO_MESTRE_PRODUTOS}'):")
                unique_produtos_coloridos = [dict(t) for t in {tuple(d.items()) for d in produtos_destacados_info}]
                for prod_info in sorted(unique_produtos_coloridos, key=lambda x: x['ID']):
                    print(f"  - ID: {prod_info['ID']}, Nome: {str(prod_info['Nome'])}")
            elif ids_produtos_oferta_set:
                print(f"ℹ️ Havia {len(ids_produtos_oferta_set)} produtos nas ofertas XML, mas nenhum correspondente foi encontrado ou pode ser colorido em '{NOME_ARQUIVO_MESTRE_PRODUTOS}'.")
            else:
                print(f"ℹ️ Nenhum produto das ofertas XML para colorir em '{NOME_ARQUIVO_MESTRE_PRODUTOS}'.")

    except Exception as e:
        print(f"❌ Erro crítico durante o gerenciamento do arquivo mestre: {e}")
        traceback.print_exc()

# ------------------------------------------
# MÓDULO DE GERAÇÃO DO RELATÓRIO FINAL DE OFERTAS
# ------------------------------------------


def carregar_dados_mestre_para_relatorio(caminho_arquivo_mestre: str) -> Optional[pd.DataFrame]:
    try:
        df_mestre = pd.read_excel(caminho_arquivo_mestre)
        colunas_esperadas = ['ID', 'SESSÃO', 'NOME_SISTEMA', 'NOME_CORRIGIDO']
        if not all(col in df_mestre.columns for col in colunas_esperadas):
            colunas_faltantes = [col for col in colunas_esperadas if col not in df_mestre.columns]
            print(f"❌ Arquivo '{NOME_ARQUIVO_MESTRE_PRODUTOS}' não contém todas as colunas esperadas. Faltando: {colunas_faltantes}. Não é possível gerar relatório.")
            return None
        return df_mestre[['ID', 'SESSÃO', 'NOME_CORRIGIDO']].copy()
    except FileNotFoundError:
        print(f"❌ Arquivo de cadastro '{NOME_ARQUIVO_MESTRE_PRODUTOS}' não encontrado. Execute a atualização do cadastro primeiro.")
        return None
    except Exception as e:
        print(f"❌ Erro ao carregar '{NOME_ARQUIVO_MESTRE_PRODUTOS}' para o relatório: {e}")
        return None


def preparar_dados_ofertas_xml_para_fusao(df_ofertas_xml_bruto: pd.DataFrame) -> pd.DataFrame:
    colunas_requeridas_xml = ['ID_Produto_XML', 'Promoção_XML', 'Preco_Promocao_XML', 'Nome_Produto_XML']
    if df_ofertas_xml_bruto.empty:
        return pd.DataFrame(columns=colunas_requeridas_xml)
    if not all(col in df_ofertas_xml_bruto.columns for col in colunas_requeridas_xml):
        print(f"❌ DataFrame de ofertas XML não contém todas as colunas requeridas: {colunas_requeridas_xml}.")
        return pd.DataFrame(columns=colunas_requeridas_xml)

    df_copia = df_ofertas_xml_bruto[colunas_requeridas_xml].copy()
    df_copia['ID_Produto_XML'] = pd.to_numeric(df_copia['ID_Produto_XML'], errors='coerce').astype('Int64')
    return df_copia


def fundir_dados_ofertas_com_mestre(df_ofertas_preparado: pd.DataFrame, df_mestre_dados: pd.DataFrame) -> pd.DataFrame:
    if df_ofertas_preparado.empty:
        colunas_esperadas_fusao = ['ID', 'NOME_PROMOÇÃO_ORIGINAL', 'PROMOÇÃO', 'Nome_Produto_XML', 'SESSÃO', 'NOME_CORRIGIDO']
        return pd.DataFrame(columns=colunas_esperadas_fusao)

    df_mestre_dados['ID'] = pd.to_numeric(df_mestre_dados['ID'], errors='coerce').astype('Int64')
    df_fundido = pd.merge(
        df_ofertas_preparado,
        df_mestre_dados,
        left_on='ID_Produto_XML',
        right_on='ID',
        how='left'
    )
    df_fundido = df_fundido.rename(columns={
        'Promoção_XML': 'NOME_PROMOÇÃO_ORIGINAL',
        'Preco_Promocao_XML': 'PROMOÇÃO',
        'ID_Produto_XML': 'ID_FINAL'
    })

    if 'ID' in df_fundido.columns and 'ID_FINAL' in df_fundido.columns:
        df_fundido['ID'] = df_fundido['ID'].fillna(df_fundido['ID_FINAL'])
    elif 'ID_FINAL' in df_fundido.columns:
        df_fundido.rename(columns={'ID_FINAL': 'ID'}, inplace=True)

    df_fundido['NOME_CORRIGIDO'] = df_fundido['NOME_CORRIGIDO'].fillna(df_fundido['Nome_Produto_XML'])
    df_fundido['SESSÃO'] = df_fundido['SESSÃO'].fillna('SEÇÃO NÃO ESPECIFICADA')
    df_fundido['NOME_PROMOÇÃO_ORIGINAL'] = df_fundido['NOME_PROMOÇÃO_ORIGINAL'].fillna('')
    return df_fundido


def enriquecer_dados_fundidos_com_detalhes_produto(df_fundido: pd.DataFrame) -> pd.DataFrame:
    if df_fundido.empty:
        return df_fundido.assign(Produto_Base_Agrupamento=pd.Series(dtype=str),
                                 Tipo_Produto_Variante=pd.Series(dtype=str),
                                 Excecao_Agrupamento=pd.Series(dtype=bool))

    df_fundido['Produto_Base_Agrupamento'] = df_fundido['NOME_CORRIGIDO'].apply(lambda x: extrair_nome_base_produto(str(x)))
    df_fundido['Tipo_Produto_Variante'] = df_fundido.apply(lambda row: extrair_tipo_variacao_produto(str(row['NOME_CORRIGIDO']), str(row['Produto_Base_Agrupamento'])), axis=1)
    df_fundido['Excecao_Agrupamento'] = df_fundido['SESSÃO'].isin(SECOES_EXCECAO_AGRUPAMENTO)
    return df_fundido


def _formatar_ofertas_para_relatorio(df_ofertas: pd.DataFrame, nome_produto_col: str, tipo_col_valor: Any) -> pd.DataFrame:
    return pd.DataFrame({
        'NOME_PROMOÇÃO': df_ofertas['NOME_PROMOÇÃO_ORIGINAL'],
        'SESSÃO': df_ofertas['SESSÃO'],
        'ID': df_ofertas['ID'],
        'PRODUTO': df_ofertas[nome_produto_col],
        'TIPO': tipo_col_valor if not isinstance(tipo_col_valor, pd.Series) else tipo_col_valor,
        'PROMOÇÃO': df_ofertas['PROMOÇÃO'],
        'DESTAQUE': False
    })


def agregar_e_transformar_ofertas_para_relatorio(df_enriquecido: pd.DataFrame) -> pd.DataFrame:
    if df_enriquecido.empty:
        return pd.DataFrame(columns=['NOME_PROMOÇÃO', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO', 'DESTAQUE'])

    lista_dfs_processados_relatorio = []

    df_excecoes = df_enriquecido[df_enriquecido['Excecao_Agrupamento']].copy()
    if not df_excecoes.empty:
        df_formatado_excecoes = _formatar_ofertas_para_relatorio(df_excecoes, 'NOME_CORRIGIDO', "")
        lista_dfs_processados_relatorio.append(df_formatado_excecoes)

    df_normais = df_enriquecido[~df_enriquecido['Excecao_Agrupamento']].copy()
    if not df_normais.empty:
        df_normais['Contagem_No_Grupo'] = df_normais.groupby(
            ['Produto_Base_Agrupamento', 'PROMOÇÃO']
        )['ID'].transform('size')

        df_para_agregar_tipos = df_normais[df_normais['Contagem_No_Grupo'] > 1].copy()
        if not df_para_agregar_tipos.empty:
            def agregar_tipos_validos(series_tipos):
                tipos_validos = [str(t).strip() for t in series_tipos if pd.notna(t) and str(t).strip()]
                return ', '.join(sorted(list(set(tipos_validos)))) if tipos_validos else ""

            df_agregado_com_tipos = df_para_agregar_tipos.groupby(
                ['Produto_Base_Agrupamento', 'PROMOÇÃO'], as_index=False
            ).agg(
                TIPO_AGREGADO=pd.NamedAgg(column='Tipo_Produto_Variante', aggfunc=agregar_tipos_validos),
                NOME_PROMOÇÃO_ORIGINAL=pd.NamedAgg(column='NOME_PROMOÇÃO_ORIGINAL', aggfunc='first'),
                ID=pd.NamedAgg(column='ID', aggfunc='first'),
                SESSÃO=pd.NamedAgg(column='SESSÃO', aggfunc='first')
            ).rename(columns={'Produto_Base_Agrupamento': 'PRODUTO', 'TIPO_AGREGADO': 'TIPO'})

            df_agregado_com_tipos['DESTAQUE'] = False
            lista_dfs_processados_relatorio.append(df_agregado_com_tipos[['NOME_PROMOÇÃO_ORIGINAL', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO', 'DESTAQUE']].rename(columns={'NOME_PROMOÇÃO_ORIGINAL': 'NOME_PROMOÇÃO'}))

        df_unicos_no_grupo = df_normais[df_normais['Contagem_No_Grupo'] == 1].copy()
        if not df_unicos_no_grupo.empty:
            df_formatado_unicos = _formatar_ofertas_para_relatorio(df_unicos_no_grupo, 'NOME_CORRIGIDO', "")
            lista_dfs_processados_relatorio.append(df_formatado_unicos)

    if not lista_dfs_processados_relatorio:
        return pd.DataFrame(columns=['NOME_PROMOÇÃO', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO', 'DESTAQUE'])

    df_relatorio_bruto_concatenado = pd.concat(lista_dfs_processados_relatorio, ignore_index=True)
    return df_relatorio_bruto_concatenado


def carregar_dados_relatorio_anterior_para_comparacao(caminho_relatorio_final: str) -> Tuple[Optional[pd.DataFrame], Set[str], Dict[str, float]]:
    try:
        df_antigo = pd.read_excel(
            caminho_relatorio_final,
            sheet_name='Ofertas Finais',
            dtype={'PRODUTO': str, 'TIPO': str, 'SESSÃO': str, 'NOME_PROMOÇÃO': str, 'ID': 'Int64'}
        )
        df_antigo['PRODUTO'] = df_antigo['PRODUTO'].fillna('')
        df_antigo['TIPO'] = df_antigo['TIPO'].fillna('')
        df_antigo['PROMOÇÃO'] = pd.to_numeric(df_antigo['PROMOÇÃO'], errors='coerce')

        if 'ID' in df_antigo.columns and df_antigo['ID'].notna().any():
            df_antigo['CHAVE_COMP'] = df_antigo['ID'].astype(str)
        else:
            df_antigo['CHAVE_COMP'] = df_antigo['PRODUTO'].astype(str) + "||" + df_antigo['TIPO'].astype(str)

        mapa_precos = pd.Series(df_antigo['PROMOÇÃO'].values, index=df_antigo['CHAVE_COMP']).to_dict()
        chaves_presentes = set(mapa_precos.keys())
        return df_antigo, chaves_presentes, mapa_precos
    except FileNotFoundError:
        print(f"ℹ️  Relatório anterior '{NOME_RELATORIO_FINAL_OFERTAS}' não encontrado. Todas as ofertas atuais serão consideradas novas.")
        return None, set(), {}
    except Exception as e:
        print(f"❌ Erro ao ler relatório anterior '{NOME_RELATORIO_FINAL_OFERTAS}': {e}. Criando novo sem destaques de atualização.")
        return None, set(), {}


def definir_destaque_para_ofertas_alteradas_ou_novas(df_relatorio_atual: pd.DataFrame, chaves_antigas_relatorio: Set[str], mapa_precos_antigos_relatorio: Dict[str, float]) -> pd.DataFrame:
    if df_relatorio_atual.empty:
        return df_relatorio_atual

    df_com_destaque = df_relatorio_atual.copy()
    if 'ID' in df_com_destaque.columns and df_com_destaque['ID'].notna().any():
        df_com_destaque['CHAVE_COMP_ATUAL'] = df_com_destaque['ID'].astype(str)
    else:
        df_com_destaque['CHAVE_COMP_ATUAL'] = df_com_destaque['PRODUTO'].astype(str) + "||" + df_com_destaque['TIPO'].astype(str)

    for indice, linha_atual in df_com_destaque.iterrows():
        chave_corrente = linha_atual['CHAVE_COMP_ATUAL']
        preco_corrente = linha_atual['PROMOÇÃO']
        marcar_como_destaque = False

        if chave_corrente not in chaves_antigas_relatorio:
            marcar_como_destaque = True
        else:
            preco_anterior = mapa_precos_antigos_relatorio.get(chave_corrente)
            if pd.isna(preco_corrente) != pd.isna(preco_anterior):
                marcar_como_destaque = True
            elif not pd.isna(preco_corrente) and not pd.isna(preco_anterior) and not np.isclose(preco_corrente, preco_anterior, equal_nan=False):
                marcar_como_destaque = True

        if marcar_como_destaque:
            df_com_destaque.loc[indice, 'DESTAQUE'] = True

    return df_com_destaque.drop(columns=['CHAVE_COMP_ATUAL'], errors='ignore')


def ordenar_dados_relatorio_e_adicionar_espacamento(df_relatorio_base: pd.DataFrame) -> pd.DataFrame:
    if df_relatorio_base.empty:
        return df_relatorio_base

    df_copia_ordenacao = df_relatorio_base.copy()
    colunas_texto_para_fillna = ['NOME_PROMOÇÃO', 'SESSÃO', 'PRODUTO', 'TIPO']
    for col in colunas_texto_para_fillna:
        if col in df_copia_ordenacao.columns:
            df_copia_ordenacao[col] = df_copia_ordenacao[col].fillna("")

    if 'ID' in df_copia_ordenacao.columns:
        df_copia_ordenacao['ID'] = pd.to_numeric(df_copia_ordenacao['ID'], errors='coerce').astype('Int64')

    if 'SESSÃO' in df_copia_ordenacao.columns:
        df_copia_ordenacao['Chave_Grupo_Secao_Ordenacao'] = df_copia_ordenacao['SESSÃO'].apply(obter_chave_ordenacao_secao)
    else:
        df_copia_ordenacao['Chave_Grupo_Secao_Ordenacao'] = np.nan

    colunas_ordenacao_lista = [COLUNA_ORDENACAO_PRIMARIA_RELATORIO]
    ordens_ascendente_lista = [ORDEM_ASCENDENTE_PRIMARIA_RELATORIO]

    chaves_secundarias_padrao = ['Chave_Grupo_Secao_Ordenacao', 'SESSÃO', 'PRODUTO']
    for chave_sec in chaves_secundarias_padrao:
        if chave_sec != COLUNA_ORDENACAO_PRIMARIA_RELATORIO and chave_sec in df_copia_ordenacao.columns:
            colunas_ordenacao_lista.append(chave_sec)
            ordens_ascendente_lista.append(True)

    if COLUNA_ORDENACAO_PRIMARIA_RELATORIO not in df_copia_ordenacao.columns:
        print(f"⚠️ Aviso: A coluna de ordenação primária '{COLUNA_ORDENACAO_PRIMARIA_RELATORIO}' não existe no DataFrame. Usando 'NOME_PROMOÇÃO' como padrão.")
        if 'NOME_PROMOÇÃO' in df_copia_ordenacao.columns:
            colunas_ordenacao_lista = ['NOME_PROMOÇÃO']
            ordens_ascendente_lista = [True]
            for chave_sec in chaves_secundarias_padrao:
                if chave_sec != 'NOME_PROMOÇÃO' and chave_sec in df_copia_ordenacao.columns:
                    colunas_ordenacao_lista.append(chave_sec)
                    ordens_ascendente_lista.append(True)
        else:
            print(f"❌ Erro: Nenhuma coluna de ordenação primária válida ('{COLUNA_ORDENACAO_PRIMARIA_RELATORIO}' ou 'NOME_PROMOÇÃO') encontrada. O relatório pode não ser ordenado corretamente.")
            colunas_ordenacao_lista = [c for c in chaves_secundarias_padrao if c in df_copia_ordenacao.columns]
            ordens_ascendente_lista = [True] * len(colunas_ordenacao_lista)

    if not colunas_ordenacao_lista:
        print("⚠️ Nenhuma coluna de ordenação válida pôde ser determinada. O relatório não será ordenado.")
        df_ordenado_final = df_copia_ordenacao.reset_index(drop=True)
    else:
        df_ordenado_final = df_copia_ordenacao.sort_values(
            by=colunas_ordenacao_lista,
            ascending=ordens_ascendente_lista
        ).reset_index(drop=True)

    lista_linhas_com_espacos_df = []
    ultimo_valor_grupo_primario_para_espaco_visto = None
    ultimo_valor_grupo_secundario_para_espaco_visto = None

    colunas_df_ordenado_com_extras = df_ordenado_final.columns.tolist()
    coluna_primaria_de_ordenacao_efetiva = colunas_ordenacao_lista[0] if colunas_ordenacao_lista else None

    if not df_ordenado_final.empty:
        for _, linha_dados in df_ordenado_final.iterrows():

            valor_grupo_primario_atual_para_espaco = None
            valor_grupo_secundario_atual_para_espaco = None

            if coluna_primaria_de_ordenacao_efetiva == 'NOME_PROMOÇÃO':
                valor_grupo_primario_atual_para_espaco = linha_dados[coluna_primaria_de_ordenacao_efetiva]
                valor_grupo_secundario_atual_para_espaco = linha_dados.get('Chave_Grupo_Secao_Ordenacao', np.nan)
            elif coluna_primaria_de_ordenacao_efetiva == 'SESSÃO':
                valor_grupo_primario_atual_para_espaco = linha_dados.get('Chave_Grupo_Secao_Ordenacao', np.nan)
            elif coluna_primaria_de_ordenacao_efetiva:
                valor_grupo_primario_atual_para_espaco = linha_dados[coluna_primaria_de_ordenacao_efetiva]
            else:
                valor_grupo_primario_atual_para_espaco = 0

            inserir_espaco_antes = False
            if ultimo_valor_grupo_primario_para_espaco_visto is not None:
                if valor_grupo_primario_atual_para_espaco != ultimo_valor_grupo_primario_para_espaco_visto:
                    inserir_espaco_antes = True
                elif coluna_primaria_de_ordenacao_efetiva == 'NOME_PROMOÇÃO' and \
                        valor_grupo_secundario_atual_para_espaco != ultimo_valor_grupo_secundario_para_espaco_visto:
                    inserir_espaco_antes = True

            if inserir_espaco_antes:
                linha_espacadora_dict = {col: np.nan for col in colunas_df_ordenado_com_extras}
                if 'DESTAQUE' in linha_espacadora_dict:
                    linha_espacadora_dict['DESTAQUE'] = False
                lista_linhas_com_espacos_df.append(linha_espacadora_dict)

            lista_linhas_com_espacos_df.append(linha_dados.to_dict())

            ultimo_valor_grupo_primario_para_espaco_visto = valor_grupo_primario_atual_para_espaco
            if coluna_primaria_de_ordenacao_efetiva == 'NOME_PROMOÇÃO':
                ultimo_valor_grupo_secundario_para_espaco_visto = valor_grupo_secundario_atual_para_espaco

    if not lista_linhas_com_espacos_df:
        df_resultado_com_espacos = df_ordenado_final
    else:
        df_resultado_com_espacos = pd.DataFrame(lista_linhas_com_espacos_df)

    if 'Chave_Grupo_Secao_Ordenacao' in df_resultado_com_espacos.columns:
        df_resultado_com_espacos = df_resultado_com_espacos.drop(columns=['Chave_Grupo_Secao_Ordenacao'], errors='ignore')

    return df_resultado_com_espacos


def _estilizar_cabecalho_planilha(planilha_ws, colunas_excel_relatorio: List[str], config_formatacao_geral: Dict, borda_padrao_celula: Border):
    linha_cabecalho_num = 1
    config_padrao = config_formatacao_geral.get('_DEFAULT_', {})
    for idx_coluna_base_zero, nome_coluna in enumerate(colunas_excel_relatorio):
        celula = planilha_ws.cell(row=linha_cabecalho_num, column=idx_coluna_base_zero + 1)
        config_col_especifica = config_formatacao_geral.get(nome_coluna, {})

        nome_fonte = config_col_especifica.get('header_font_name', config_padrao.get('header_font_name', 'Calibri'))
        tamanho_fonte = config_col_especifica.get('header_font_size', config_padrao.get('header_font_size', 9))
        cor_fonte_hex = config_col_especifica.get('header_font_color', config_padrao.get('header_font_color', "000000")).replace("#", "")
        negrito_fonte = config_col_especifica.get('header_font_bold', config_padrao.get('header_font_bold', True))
        italico_fonte = config_col_especifica.get('header_font_italic', config_padrao.get('header_font_italic', False))
        cor_fundo_hex = config_col_especifica.get('header_fill_color', config_padrao.get('header_fill_color', None))
        if cor_fundo_hex:
            cor_fundo_hex = cor_fundo_hex.replace("#", "")
        align_h = config_col_especifica.get('header_alignment_horizontal', config_padrao.get('header_alignment_horizontal', 'left'))
        align_v = config_col_especifica.get('header_alignment_vertical', config_padrao.get('header_alignment_vertical', 'center'))

        celula.font = Font(name=nome_fonte, size=tamanho_fonte, color=cor_fonte_hex, bold=negrito_fonte, italic=italico_fonte)
        if cor_fundo_hex:
            celula.fill = PatternFill(start_color=cor_fundo_hex, end_color=cor_fundo_hex, fill_type="solid")
        celula.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=False)
        celula.border = borda_padrao_celula


def _estilizar_linhas_de_dados_planilha(planilha_ws, df_completo_com_destaque_info: pd.DataFrame, colunas_excel_relatorio: List[str], config_formatacao_geral: Dict, borda_padrao_celula: Border, preenchimento_linha_destacada: PatternFill, fonte_linha_destacada: Font):
    config_padrao = config_formatacao_geral.get('_DEFAULT_', {})
    for idx_linha_df, (_, linha_df_com_destaque) in enumerate(df_completo_com_destaque_info.iterrows()):
        num_linha_excel = idx_linha_df + 2

        eh_linha_espacadora = all(pd.isna(val) for col, val in linha_df_com_destaque.items() if col != 'DESTAQUE')
        if eh_linha_espacadora:
            continue

        destacar_esta_linha = linha_df_com_destaque.get('DESTAQUE', False)

        for idx_coluna_base_zero, nome_coluna_excel in enumerate(colunas_excel_relatorio):
            celula = planilha_ws.cell(row=num_linha_excel, column=idx_coluna_base_zero + 1)
            config_col_especifica = config_formatacao_geral.get(nome_coluna_excel, {})

            nome_fonte_base = config_col_especifica.get('font_name', config_padrao.get('font_name', 'Calibri'))
            tamanho_fonte_base = config_col_especifica.get('font_size', config_padrao.get('font_size', 10))
            cor_fonte_base_hex = config_col_especifica.get('font_color', config_padrao.get('font_color', "000000")).replace("#", "")
            negrito_fonte_base = config_col_especifica.get('font_bold', config_padrao.get('font_bold', False))
            italico_fonte_base = config_col_especifica.get('font_italic', config_padrao.get('font_italic', False))
            cor_fundo_base_hex = config_col_especifica.get('fill_color', config_padrao.get('fill_color', None))
            if cor_fundo_base_hex:
                cor_fundo_base_hex = cor_fundo_base_hex.replace("#", "")

            formato_numero = config_col_especifica.get('number_format', config_padrao.get('number_format', '@'))
            align_h = config_col_especifica.get('alignment_horizontal', config_padrao.get('alignment_horizontal', 'left'))
            align_v = config_col_especifica.get('alignment_vertical', config_padrao.get('alignment_vertical', 'center'))

            fonte_final_celula = Font(name=nome_fonte_base, size=tamanho_fonte_base, color=cor_fonte_base_hex, bold=negrito_fonte_base, italic=italico_fonte_base)
            preenchimento_final_celula = PatternFill(fill_type=None)
            if cor_fundo_base_hex:
                preenchimento_final_celula = PatternFill(start_color=cor_fundo_base_hex, end_color=cor_fundo_base_hex, fill_type="solid")

            if destacar_esta_linha:
                preenchimento_final_celula = preenchimento_linha_destacada
                fonte_final_celula = Font(
                    name=fonte_linha_destacada.name if fonte_linha_destacada.name else nome_fonte_base,
                    size=fonte_linha_destacada.size if fonte_linha_destacada.size else tamanho_fonte_base,
                    color=fonte_linha_destacada.color.rgb if fonte_linha_destacada.color else cor_fonte_base_hex,
                    bold=fonte_linha_destacada.bold if fonte_linha_destacada.bold is not None else negrito_fonte_base,
                    italic=fonte_linha_destacada.italic if fonte_linha_destacada.italic is not None else italico_fonte_base
                )

            celula.font = fonte_final_celula
            celula.fill = preenchimento_final_celula
            celula.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=False)
            celula.border = borda_padrao_celula

            if celula.value is not None:
                if formato_numero not in ['@', 'General'] and isinstance(celula.value, str):
                    try:
                        celula.value = float(str(celula.value).replace("R$", "").replace(".", "").replace(",", ".").strip())
                    except ValueError:
                        pass

                if formato_numero == 'R$ #,##0.00' and isinstance(celula.value, (int, float)) and celula.value == 0:
                    celula.value = 0.00
                celula.number_format = formato_numero


def _ajustar_largura_das_colunas_planilha(planilha_ws, colunas_excel_relatorio: List[str], config_formatacao_geral: Dict):
    config_padrao = config_formatacao_geral.get('_DEFAULT_', {})
    for idx_coluna_base_zero, nome_coluna_excel in enumerate(colunas_excel_relatorio):
        letra_coluna = get_column_letter(idx_coluna_base_zero + 1)
        comprimento_maximo_conteudo = 0

        config_col_especifica = config_formatacao_geral.get(nome_coluna_excel, {})
        formato_numero_atual = config_col_especifica.get('number_format', config_padrao.get('number_format', '@'))

        for num_linha_ws in range(1, planilha_ws.max_row + 1):
            celula = planilha_ws.cell(row=num_linha_ws, column=idx_coluna_base_zero + 1)
            if celula.value is not None:
                valor_texto_celula = str(celula.value)
                if isinstance(celula.value, (int, float)) and formato_numero_atual == 'R$ #,##0.00':
                    valor_texto_celula = f"R$ {celula.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                elif isinstance(celula.value, (int, float)) and formato_numero_atual == '0':
                    valor_texto_celula = f"{celula.value:.0f}"

                comprimento_atual_conteudo = len(valor_texto_celula)
                comprimento_maximo_conteudo = max(comprimento_maximo_conteudo, comprimento_atual_conteudo)

        valor_cabecalho_celula = planilha_ws.cell(row=1, column=idx_coluna_base_zero + 1).value
        if valor_cabecalho_celula:
            comprimento_maximo_conteudo = max(comprimento_maximo_conteudo, len(str(valor_cabecalho_celula)))

        largura_ajustada = min(comprimento_maximo_conteudo + 4, 60)
        planilha_ws.column_dimensions[letra_coluna].width = largura_ajustada if largura_ajustada > 8 else 10


def escrever_e_estilizar_planilha_relatorio_final(df_dados_para_planilha: pd.DataFrame, df_com_info_destaque_completa: pd.DataFrame, caminho_relatorio_final: str, colunas_finais_para_excel: List[str], config_formatacao: Dict, preenchimento_destaque_excel: PatternFill, fonte_destaque_excel: Font) -> None:
    try:
        with pd.ExcelWriter(caminho_relatorio_final, engine='openpyxl') as excel_writer:
            df_para_salvar_na_planilha = df_dados_para_planilha[colunas_finais_para_excel].copy()
            df_para_salvar_na_planilha.to_excel(excel_writer, index=False, sheet_name='Ofertas Finais')
            planilha_ativa_ws = excel_writer.sheets['Ofertas Finais']

            planilha_ativa_ws.sheet_view.showGridLines = False
            borda_fina_lateral = Side(style='thin')
            estilo_borda_completa = Border(left=borda_fina_lateral, right=borda_fina_lateral, top=borda_fina_lateral, bottom=borda_fina_lateral)

            _estilizar_cabecalho_planilha(planilha_ativa_ws, colunas_finais_para_excel, config_formatacao, estilo_borda_completa)
            _estilizar_linhas_de_dados_planilha(planilha_ativa_ws, df_com_info_destaque_completa, colunas_finais_para_excel, config_formatacao, estilo_borda_completa, preenchimento_destaque_excel, fonte_destaque_excel)
            _ajustar_largura_das_colunas_planilha(planilha_ativa_ws, colunas_finais_para_excel, config_formatacao)

        num_ofertas_reais_no_relatorio = 0
        if 'PRODUTO' in df_dados_para_planilha.columns:
            num_ofertas_reais_no_relatorio = df_dados_para_planilha['PRODUTO'].notna().sum()

        print(f"✅ Relatório '{NOME_RELATORIO_FINAL_OFERTAS}' criado com {num_ofertas_reais_no_relatorio} ofertas (incluindo linhas de dados, excluindo espaçadores) na raiz do projeto.")

    except Exception as e:
        print(f"❌ Erro ao salvar/formatar Excel '{NOME_RELATORIO_FINAL_OFERTAS}': {e}")
        traceback.print_exc()


def gerar_relatorio_consolidado_ofertas(df_ofertas_consolidadas_xml: pd.DataFrame, modo_operacao: str) -> None:
    print(f"📊 Gerando o relatório final de ofertas: './{NOME_RELATORIO_FINAL_OFERTAS}'...")
    caminho_mestre_prod = f'./{NOME_ARQUIVO_MESTRE_PRODUTOS}'
    caminho_relatorio = f'./{NOME_RELATORIO_FINAL_OFERTAS}'
    colunas_excel_finais_ordenadas = ['NOME_PROMOÇÃO', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO']

    df_mestre_para_relatorio = carregar_dados_mestre_para_relatorio(caminho_mestre_prod)
    if df_mestre_para_relatorio is None:
        return

    if df_ofertas_consolidadas_xml.empty and modo_operacao == 'create':
        print(f"⚠️ DataFrame de ofertas XML está vazio. Nada a processar para '{NOME_RELATORIO_FINAL_OFERTAS}' no modo 'create'.")
        return
    elif df_ofertas_consolidadas_xml.empty and modo_operacao == 'update':
        print(f"⚠️ DataFrame de ofertas XML está vazio. Tentando gerar relatório '{NOME_RELATORIO_FINAL_OFERTAS}' para limpar/atualizar com base no anterior.")

    df_ofertas_xml_preparado = preparar_dados_ofertas_xml_para_fusao(df_ofertas_consolidadas_xml)

    df_fundido_ofertas_mestre = fundir_dados_ofertas_com_mestre(df_ofertas_xml_preparado, df_mestre_para_relatorio)
    df_enriquecido_com_detalhes = enriquecer_dados_fundidos_com_detalhes_produto(df_fundido_ofertas_mestre)
    df_relatorio_processado_agregado = agregar_e_transformar_ofertas_para_relatorio(df_enriquecido_com_detalhes)

    for col_excel in colunas_excel_finais_ordenadas:
        if col_excel not in df_relatorio_processado_agregado.columns:
            df_relatorio_processado_agregado[col_excel] = np.nan if col_excel in ['ID', 'PROMOÇÃO'] else ""
    if 'DESTAQUE' not in df_relatorio_processado_agregado.columns:
        df_relatorio_processado_agregado['DESTAQUE'] = False
    df_relatorio_processado_agregado['DESTAQUE'] = df_relatorio_processado_agregado['DESTAQUE'].fillna(False).astype(bool)

    df_antigo_relatorio_comp, chaves_antigas_comp, mapa_precos_antigos_comp = None, set(), {}
    if modo_operacao == 'update':
        df_antigo_relatorio_comp, chaves_antigas_comp, mapa_precos_antigos_comp = carregar_dados_relatorio_anterior_para_comparacao(caminho_relatorio)
        df_relatorio_processado_agregado = definir_destaque_para_ofertas_alteradas_ou_novas(df_relatorio_processado_agregado, chaves_antigas_comp, mapa_precos_antigos_comp)

    df_relatorio_final_ordenado_com_espacos = ordenar_dados_relatorio_e_adicionar_espacamento(df_relatorio_processado_agregado)

    if df_relatorio_final_ordenado_com_espacos.empty:
        print(f"⚠️ Nenhum dado para gerar o relatório '{NOME_RELATORIO_FINAL_OFERTAS}'.")
        if modo_operacao == 'update' and os.path.exists(caminho_relatorio):
            try:
                pd.DataFrame(columns=colunas_excel_finais_ordenadas).to_excel(caminho_relatorio, index=False, sheet_name='Ofertas Finais')
                print(f"ℹ️ Relatório anterior '{NOME_RELATORIO_FINAL_OFERTAS}' substituído por um arquivo vazio (sem ofertas atuais).")
            except Exception as e_salvar_vazio:
                print(f"❌ Erro ao tentar salvar relatório vazio: {e_salvar_vazio}")
        return

    escrever_e_estilizar_planilha_relatorio_final(
        df_relatorio_final_ordenado_com_espacos,
        df_relatorio_final_ordenado_com_espacos,
        caminho_relatorio,
        colunas_excel_finais_ordenadas,
        CONFIGURACAO_FORMATACAO_COLUNAS,
        PREENCHIMENTO_DESTAQUE,
        FONTE_DESTAQUE
    )

    if modo_operacao == 'update':
        num_destacados = df_relatorio_processado_agregado['DESTAQUE'].sum() if 'DESTAQUE' in df_relatorio_processado_agregado.columns else 0
        num_ofertas_reais = len(df_relatorio_processado_agregado.dropna(subset=['PRODUTO'], how='all'))

        if df_antigo_relatorio_comp is not None:
            print(f"   ✨ {num_destacados} ofertas foram destacadas como novas ou com preço alterado.")
        elif not chaves_antigas_comp and num_ofertas_reais > 0:
            print(f"   ✨ Todas as {num_ofertas_reais} ofertas são consideradas novas (nenhum relatório anterior para comparar).")


# ------------------------------------------
# FUNÇÃO PRINCIPAL DE ORQUESTRAÇÃO
# ------------------------------------------
def executar_processamento_principal(modo_operacao: str) -> None:
    print(f"🚀 Iniciando processamento de ofertas (Modo: {modo_operacao})...")

    if not os.path.isdir(DIRETORIO_ENTRADA_XML):
        print(f"❌ ERRO: Diretório de entrada '{DIRETORIO_ENTRADA_XML}' não encontrado. Crie-o e adicione os arquivos XML.")
        print("🏁 Processamento abortado.")
        return

    arquivos_xml_originais_encontrados = listar_arquivos_por_extensao(DIRETORIO_ENTRADA_XML, EXTENSAO_XML)
    if not arquivos_xml_originais_encontrados:
        print(f"ℹ️ Nenhum arquivo XML encontrado em '{DIRETORIO_ENTRADA_XML}'.")
    else:
        normalizar_nomes_arquivos_xml_em_lote(arquivos_xml_originais_encontrados, DIRETORIO_ENTRADA_XML)

    arquivos_xml_efetivos_para_processar = [
        f for f in os.listdir(DIRETORIO_ENTRADA_XML)
        if f.endswith(EXTENSAO_XML) and os.path.isfile(os.path.join(DIRETORIO_ENTRADA_XML, f))
    ]
    if not arquivos_xml_efetivos_para_processar:
        print(f"ℹ️ Nenhum arquivo XML para processar em '{DIRETORIO_ENTRADA_XML}' após normalização.")

    df_ofertas_consolidadas_xml_geral = consolidar_dados_de_multiplos_arquivos_xml(arquivos_xml_efetivos_para_processar, DIRETORIO_ENTRADA_XML)

    if df_ofertas_consolidadas_xml_geral.empty and not arquivos_xml_efetivos_para_processar:
        print("ℹ️ Nenhum dado de produto foi consolidado dos arquivos XML nesta execução.")

    gerenciar_atualizacao_arquivo_mestre_produtos(df_ofertas_consolidadas_xml_geral)

    necessario_gerar_relatorio = True
    caminho_relatorio_final_para_verificacao = f'./{NOME_RELATORIO_FINAL_OFERTAS}'
    if df_ofertas_consolidadas_xml_geral.empty:
        if modo_operacao == 'create':
            print("🏁 Processamento concluído (sem dados XML para gerar relatório de ofertas no modo 'create').")
            necessario_gerar_relatorio = False
        elif modo_operacao == 'update' and not os.path.exists(caminho_relatorio_final_para_verificacao):
            print("🏁 Processamento concluído (sem dados XML e sem relatório anterior para atualizar).")
            necessario_gerar_relatorio = False

    if necessario_gerar_relatorio:
        gerar_relatorio_consolidado_ofertas(df_ofertas_consolidadas_xml_geral, modo_operacao)

    print("🏁 Processamento concluído.")


# ------------------------------------------
# INTERFACE DE LINHA DE COMANDO (CLI)
# ------------------------------------------
if __name__ == "__main__":
    analisador_argumentos = argparse.ArgumentParser(description="Processa arquivos XML de ofertas e gera relatórios Excel.")
    analisador_argumentos.add_argument(
        "--mode",
        choices=['create', 'update'],
        default='create',
        help="Modo de operação: 'create' para gerar do zero (padrão), 'update' para destacar mudanças no relatório de ofertas."
    )
    argumentos_fornecidos = analisador_argumentos.parse_args()
    executar_processamento_principal(modo_operacao=argumentos_fornecidos.mode)
