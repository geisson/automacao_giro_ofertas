# -*- coding: utf-8 -*-
"""
Script para processamento de arquivos XML de ofertas, refatorado com paradigma funcional.
Dividido em seções bem definidas com funções puras sempre que possível.
Modificado para incluir agrupamento condicional de produtos, formatação de Excel
avançada (contábil, bordas) e ordenação específica de seções com linhas em branco.
V6: Corrigida a lógica de identificação de linhas de espaçamento para não aplicar bordas.
V7: Adicionada funcionalidade para limpar cores e destacar produtos de oferta em verde no 'produtos_cadastrados.xlsx'.
V8 (Sua Solicitação): Modificada a lógica de agrupamento no relatório 'giro_da_praça_ofertas.xlsx'
    para agrupar por nome base e preço, criando uma coluna 'TIPO' com as variantes.
V8.1: Refatorada a criação de DataFrames antes da concatenação para evitar FutureWarning.
V8.2: Refinada a lógica para garantir que produtos não agrupados (exceções, únicos por preço/nome base)
      usem o nome completo original e tenham a coluna TIPO vazia.
V8.3: Corrigido ValueError: cannot reindex on an axis with duplicate labels.
V9: Adicionada configuração granular de formatação de colunas (fonte, cor, negrito, tipo de dado, etc.)
    para 'giro_da_praça_ofertas.xlsx'.
"""

# ------------------------------------------
# DEPENDÊNCIAS
# ------------------------------------------
import pandas as pd
import os
import glob
import xmltodict
from typing import List, Dict, Any, Optional
import re
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import traceback

# ------------------------------------------
# CONSTANTES
# ------------------------------------------
INPUT_XML_DIR = './arquivos_xml_entrada/'
XML_EXTENSION = '.xml'
PRODUCT_MASTER_FILE_NAME = 'produtos_cadastrados.xlsx'
FINAL_OFFERS_REPORT_NAME = 'giro_da_praça_ofertas.xlsx'
SECOES_EXCECAO_AGRUPAMENTO = ["#01 MERCEARIA - #01 ALTO GIRO", "#01 MERCEARIA - #02 ALTO GIRO"]

# Configuração de Formatação de Colunas para giro_da_praça_ofertas.xlsx
# Cores são em formato ARGB (ex: "FF000000" para preto opaco, "000000" é o mesmo que "FF000000")
COLUMNS_FORMATTING_CONFIG = {
    '_DEFAULT_': {
        'font_name': 'Calibri',
        'font_size': 10,
        'font_color': "000000", # Preto
        'font_bold': False,
        'font_italic': False,
        'number_format': '@',    # Formato de texto por padrão
        'alignment_horizontal': 'left',
        'alignment_vertical': 'center',

        'header_font_name': 'Calibri',
        'header_font_size': 9,
        'header_font_color': "000000", # Branco
        'header_font_bold': True,
        'header_font_italic': False,
        'header_fill_color': "ffffff", # Azul Padrão do Excel
        'header_alignment_horizontal': 'left',
        'header_alignment_vertical': 'center',
    },
    'NOME_PROMOÇÃO': {
        'font_size': 7,
        # Herda outras formatações do _DEFAULT_
    },
    'SESSÃO': {
        'font_size': 7,
        'font_color': "000000", # Azul escuro
    },
    'ID': {
        'font_size': 7 ,
        'number_format': '0', # Número inteiro
    },
    'PRODUTO': {
        'font_size': 9,
    },
    'TIPO': {
        'font_size': 9,
        'font_color': "000000", # Cinza
        'font_italic': True
    },
    'PROMOÇÃO': {
        'font_size': 10,
        'font_color': "000000", # Verde escuro para promoção
        'font_bold': True,
        'number_format': 'R$ #,##0.00', # Formato Contábil Brasileiro
        'alignment_horizontal': 'right',
    }
}

# (O restante das funções de tratamento de texto, manipulação de arquivos, XML, etc. permanecem iguais)
# ...
# FUNÇÕES PARA TRATAMENTO DE TEXTO
# ... (igual ao anterior)
def remover_acentuacao(texto: str) -> str:
    if not isinstance(texto, str): return ""
    mapeamento_acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u', 'ã': 'a', 'õ': 'o', 'ä': 'a', 'ë': 'e', 'ï': 'i',
        'ö': 'o', 'ü': 'u', 'ç': 'c', 'ñ': 'n', 'ÿ': 'y', 'ý': 'y', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O',
        'Ú': 'U', 'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U', 'Â': 'A', 'Ê': 'E', 'Î': 'I', 'Ô': 'O',
        'Û': 'U', 'Ã': 'A', 'Õ': 'O', 'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U', 'Ç': 'C', 'Ñ': 'N',
        'Ÿ': 'Y', 'Ý': 'Y'
    }
    return ''.join(mapeamento_acentos.get(c, c) for c in texto)

def normalizar_texto(texto: str) -> str:
    if not isinstance(texto, str): return ""
    texto_sem_acentos = remover_acentuacao(texto)
    return texto_sem_acentos.lower()

# FUNÇÕES PARA MANIPULAÇÃO DE ARQUIVOS
# ... (igual ao anterior)
def listar_arquivos_xml(diretorio: str) -> List[str]:
    return glob.glob(os.path.join(diretorio, f'*{XML_EXTENSION}'))

def renomear_arquivo(arquivo_antigo: str, diretorio: str) -> None:
    nome_arquivo = os.path.basename(arquivo_antigo)
    novo_nome = normalizar_texto(nome_arquivo)
    novo_caminho = os.path.join(diretorio, novo_nome)
    try:
        if arquivo_antigo != novo_caminho:
            os.rename(arquivo_antigo, novo_caminho)
    except FileExistsError:
        print(f"⚠️  Arquivo '{novo_caminho}' já existe. Não foi renomeado '{arquivo_antigo}'.")
    except Exception as e:
        print(f"❌ Erro ao renomear '{arquivo_antigo}' para '{novo_caminho}': {e}")

def renomear_arquivos_em_lote(arquivos: List[str], diretorio: str) -> None:
    for arquivo in arquivos:
        renomear_arquivo(arquivo, diretorio)

# FUNÇÕES PARA PROCESSAMENTO DE XML
# ... (igual ao anterior)
def parse_produto_xml(produto_xml_data: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "Promoção_XML": produto_xml_data.get('descrpromocao', ''),
        "ID_Produto_XML": int(produto_xml_data.get('idsubproduto', 0)),
        "Nome_Produto_XML": produto_xml_data.get('descrresproduto', ''),
        "Preco_Promocao_XML": float(produto_xml_data.get('precopromocao', 0.0) if produto_xml_data.get('precopromocao') not in [None, ''] else 0.0)
    }

def processar_arquivo_xml(nome_arquivo_xml: str) -> List[Dict[str, Any]]:
    caminho_arquivo = os.path.join(INPUT_XML_DIR, nome_arquivo_xml)
    try:
        with open(caminho_arquivo, 'rb') as arquivo_xml:
            dados_xml = xmltodict.parse(arquivo_xml)
            if 'temporario_846' not in dados_xml:
                print(f"⚠️ Estrutura 'temporario_846' não encontrada em {nome_arquivo_xml}.")
                return []
            temporario_846 = dados_xml['temporario_846']
            produtos_data_xml = temporario_846.get('temporario_846_row', [])
            if isinstance(produtos_data_xml, dict): produtos_data_xml = [produtos_data_xml]
            if not produtos_data_xml: return []
            return [parse_produto_xml(p) for p in produtos_data_xml if isinstance(p, dict)]
    except FileNotFoundError: print(f"❌ Arquivo XML não encontrado: {caminho_arquivo}"); return []
    except Exception as e: print(f"❌ Erro ao processar XML {nome_arquivo_xml}: {e}"); traceback.print_exc(); return []

# FUNÇÃO PARA CONSOLIDAÇÃO DE DADOS XML
# ... (igual ao anterior)
def consolidar_dados_de_xmls(lista_arquivos_xml: List[str]) -> pd.DataFrame:
    todos_produtos_lista = []
    for arquivo_nome in lista_arquivos_xml:
        produtos_do_arquivo = processar_arquivo_xml(arquivo_nome)
        if produtos_do_arquivo:
            todos_produtos_lista.extend(produtos_do_arquivo)
            print(f"✅ {arquivo_nome}: {len(produtos_do_arquivo)} produtos processados.")
    if not todos_produtos_lista: print("⚠️ Nenhum produto encontrado nos arquivos XML."); return pd.DataFrame()
    return pd.DataFrame(todos_produtos_lista)

# FUNÇÕES AUXILIARES PARA INTEGRAÇÃO DE DADOS
# ... (igual ao anterior)
def extract_product_base_name(product_name: str) -> str:
    if not isinstance(product_name, str): return ""
    unidades = [
        "KILOGRAMAS", "KILOGRAMA", "QUILOGRAMAS", "QUILOGRAMA", "LITROS", "LITRO", "UNIDADES", "UNIDADE",
        "PACOTES", "PACOTE", "CAIXAS", "CAIXA", "FARDOS", "FARDO", "GARRAFAS", "GARRAFA", "LATAS", "LATA",
        "POTES", "POTE", "ROLOS", "ROLO", "SACHES", "SACHE", "TUBOS", "TUBO", "VIDROS", "VIDRO", "SACOLAS",
        "SACOLA", "BLISTERS", "BLISTER", "CARTELAS", "CARTELA", "DISPLAYS", "DISPLAY", "GALÕES", "GALÃO",
        "REFILs", "REFIL", "TABLETES", "TABLETE", "PEÇAS", "PEÇA", "FOLHAS", "FOLHA", "FLACONETES",
        "FLACONETE", "BISNAGAS", "BISNAGA", "AMPOLAS", "AMPOLA", "DRAGEAS", "CAPSULAS", "CAPSULA",
        "COMPRIMIDOS", "COMPRIMIDO", "ENVELOPES", "ENVELOPE",
        "KG", "GR", "G", "MG", "L", "ML", "CM", "MM", "M",
        "UN", "UND", "UNID", "PC", "PÇ", "PCT", "CX", "FD", "GF", "LT", "PT", "RL", "SC", "TB", "VD",
        "BL", "CT", "DP", "GL", "RF", "TBL", "FL", "CAPS", "COMP", "ENV"
    ]
    product_name_norm = ' '.join(product_name.split())
    unidades_regex_part = "|".join(map(re.escape, sorted(unidades, key=len, reverse=True)))
    regex_str = r"^(.*?)(\s?)(\d+[\.,]?\d*)(\s*)(" + unidades_regex_part + r")($|\s+.*$)"
    match = re.match(regex_str, product_name_norm, re.IGNORECASE)
    if match:
        nome_antes_da_medida = match.group(1).strip()
        quantidade = match.group(3)
        unidade_texto = match.group(5)
        base_nome = f"{nome_antes_da_medida} {quantidade}{unidade_texto}".strip()
        return ' '.join(base_nome.split())
    return product_name_norm

def extract_product_type_from_full_name(full_name: str, base_name: str) -> str:
    if not isinstance(full_name, str) or not isinstance(base_name, str): return ""
    full_name_norm = ' '.join(full_name.split())
    base_name_norm = ' '.join(base_name.split())
    if full_name_norm == base_name_norm: return ""
    if full_name_norm.startswith(base_name_norm):
        type_part = full_name_norm[len(base_name_norm):].strip()
        type_part = re.sub(r"^[ \-\/\_]+", "", type_part).strip()
        return type_part
    return ""

def get_section_sort_key(section_name: Any) -> str:
    if not isinstance(section_name, str) or pd.isna(section_name): return "SEM SEÇÃO DEFINIDA"
    if section_name.startswith("#01 MERCEARIA - #") and "ALTO GIRO" in section_name:
        return "#01 MERCEARIA - ALTO GIRO"
    return section_name.split(" - ", 1)[0]

def format_individual_product_for_report(df_input: pd.DataFrame, colunas_finais: list) -> pd.DataFrame:
    if df_input.empty: return pd.DataFrame(columns=colunas_finais)
    df_output = pd.DataFrame()
    df_output['NOME_PROMOÇÃO'] = df_input['Promoção_XML']
    df_output['SESSÃO'] = df_input['SESSÃO']
    df_output['ID'] = df_input['ID_Produto_XML']
    df_output['PRODUTO'] = df_input['NOME_CORRIGIDO']
    df_output['TIPO'] = ""
    df_output['PROMOÇÃO'] = df_input['Preco_Promocao_XML']
    return df_output.reindex(columns=colunas_finais)

# FUNÇÕES PARA INTEGRAÇÃO E MANUTENÇÃO DE DADOS
# ... (igual ao anterior)
def atualizar_cadastro_produtos_base(df_ofertas_xml: pd.DataFrame) -> None:
    path_cadastro_base = f'./{PRODUCT_MASTER_FILE_NAME}'
    cols_cadastro_base_ordenadas = ['ID', 'SESSÃO', 'NOME_SISTEMA', 'NOME_CORRIGIDO']
    try:
        df_cadastro_base = pd.read_excel(path_cadastro_base)
        if 'ID' not in df_cadastro_base.columns:
            print(f"⚠️  Arquivo '{PRODUCT_MASTER_FILE_NAME}' não possui coluna 'ID'. Tratando como novo.")
            df_cadastro_base = pd.DataFrame(columns=cols_cadastro_base_ordenadas)
        temp_df = pd.DataFrame(columns=cols_cadastro_base_ordenadas)
        for col in cols_cadastro_base_ordenadas:
            if col in df_cadastro_base.columns: temp_df[col] = df_cadastro_base[col]
            else: temp_df[col] = np.nan
        df_cadastro_base = temp_df
    except FileNotFoundError:
        print(f"ℹ️  Arquivo '{PRODUCT_MASTER_FILE_NAME}' não encontrado. Será criado.")
        df_cadastro_base = pd.DataFrame(columns=cols_cadastro_base_ordenadas)
    except Exception as e: print(f"❌ Erro ao carregar '{PRODUCT_MASTER_FILE_NAME}': {e}. Operação cancelada."); return

    novos_produtos_df = pd.DataFrame()
    if df_ofertas_xml.empty: print(f"ℹ️ DataFrame de ofertas XML está vazio. Nenhum produto novo para adicionar ao '{PRODUCT_MASTER_FILE_NAME}'.")
    else:
        if 'ID_Produto_XML' not in df_ofertas_xml.columns or 'Nome_Produto_XML' not in df_ofertas_xml.columns:
            print(f"❌ Colunas 'ID_Produto_XML' ou 'Nome_Produto_XML' ausentes no DataFrame de ofertas. Não é possível atualizar o cadastro.")
        else:
            if 'ID' in df_cadastro_base.columns:
                df_cadastro_base['ID'] = pd.to_numeric(df_cadastro_base['ID'], errors='coerce')
                ids_existentes_cadastro = set(df_cadastro_base['ID'].dropna().astype(int).unique())
            else: ids_existentes_cadastro = set()
            df_ofertas_xml['ID_Produto_XML'] = pd.to_numeric(df_ofertas_xml['ID_Produto_XML'], errors='coerce')
            novos_produtos_df = df_ofertas_xml[~df_ofertas_xml['ID_Produto_XML'].dropna().astype(int).isin(ids_existentes_cadastro)].copy()
            if novos_produtos_df.empty: print(f"ℹ️ Nenhum produto novo das ofertas XML para adicionar ao '{PRODUCT_MASTER_FILE_NAME}'.")
            else:
                print(f"ℹ️ {len(novos_produtos_df)} novos produtos encontrados para adicionar.")
                novos_para_cadastro_list = [{'ID': row['ID_Produto_XML'], 'NOME_SISTEMA': row['Nome_Produto_XML'], 'SESSÃO': 'SEÇÃO NÃO ESPECIFICADA', 'NOME_CORRIGIDO': row['Nome_Produto_XML']} for _, row in novos_produtos_df.iterrows()]
                df_novos_formatados = pd.DataFrame(novos_para_cadastro_list, columns=cols_cadastro_base_ordenadas)
                df_cadastro_base = pd.concat([df_cadastro_base, df_novos_formatados], ignore_index=True)
                df_cadastro_base.drop_duplicates(subset=['ID'], keep='first', inplace=True)
    try:
        df_para_salvar = df_cadastro_base.reindex(columns=cols_cadastro_base_ordenadas)
        df_para_salvar = df_para_salvar.sort_values(by=['SESSÃO', 'NOME_SISTEMA', 'ID'])
        if 'ID' in df_para_salvar.columns: df_para_salvar['ID'] = pd.to_numeric(df_para_salvar['ID'], errors='coerce')
        df_para_salvar.to_excel(path_cadastro_base, index=False)
        if not df_para_salvar.empty or os.path.exists(path_cadastro_base):
            green_fill, no_fill = PatternFill(start_color="45A045", end_color="45A045", fill_type="solid"), PatternFill(fill_type=None)
            ids_produtos_oferta_xml = set(df_ofertas_xml['ID_Produto_XML'].dropna().astype(int).unique()) if not df_ofertas_xml.empty and 'ID_Produto_XML' in df_ofertas_xml.columns else set()
            produtos_coloridos_info_list, wb = [], load_workbook(path_cadastro_base)
            ws, id_col_idx, nome_corr_idx, nome_sist_idx = wb.active, None, None, None
            if ws.max_row > 0:
                for i, cell in enumerate(ws[1]):
                    if cell.value == 'ID': id_col_idx = i + 1
                    elif cell.value == 'NOME_CORRIGIDO': nome_corr_idx = i + 1
                    elif cell.value == 'NOME_SISTEMA': nome_sist_idx = i + 1
            if id_col_idx is None and ws.max_row > 0: print(f"⚠️  Aviso: Coluna 'ID' não encontrada no cabeçalho de '{PRODUCT_MASTER_FILE_NAME}'. Não será possível aplicar cores.")
            elif ws.max_row <= 1 and not ids_produtos_oferta_xml: print(f"ℹ️  Arquivo '{PRODUCT_MASTER_FILE_NAME}' está vazio ou contém apenas cabeçalho e não há produtos XML para colorir.")
            elif ws.max_row > 1 and id_col_idx is not None:
                for r_num in range(2, ws.max_row + 1):
                    for c_num in range(1, ws.max_column + 1): ws.cell(row=r_num, column=c_num).fill = no_fill
                    id_val = ws.cell(row=r_num, column=id_col_idx).value
                    if id_val is not None:
                        try:
                            curr_id = int(float(id_val))
                            if curr_id in ids_produtos_oferta_xml:
                                for c_num in range(1, ws.max_column + 1): ws.cell(row=r_num, column=c_num).fill = green_fill
                                nome_prod = "NOME NÃO ENCONTRADO"
                                if nome_corr_idx and ws.cell(row=r_num, column=nome_corr_idx).value: nome_prod = ws.cell(row=r_num, column=nome_corr_idx).value
                                elif nome_sist_idx and ws.cell(row=r_num, column=nome_sist_idx).value: nome_prod = ws.cell(row=r_num, column=nome_sist_idx).value
                                produtos_coloridos_info_list.append({'ID': curr_id, 'Nome': nome_prod})
                        except (ValueError, TypeError): pass
            wb.save(path_cadastro_base)
            if not novos_produtos_df.empty: print(f"✅ '{PRODUCT_MASTER_FILE_NAME}' atualizado com {len(novos_produtos_df)} novos produtos e formatação de cores aplicada.")
            else: print(f"✅ '{PRODUCT_MASTER_FILE_NAME}' salvo com formatação de cores aplicada.")
            if produtos_coloridos_info_list:
                print(f"ℹ️ Produtos das ofertas XML (coloridos em verde em '{PRODUCT_MASTER_FILE_NAME}'):")
                unique_prods = [dict(t) for t in {tuple(d.items()) for d in produtos_coloridos_info_list}]
                for info in sorted(unique_prods, key=lambda x: x['ID']): print(f"  - ID: {info['ID']}, Nome: {str(info['Nome'])}")
            elif ids_produtos_oferta_xml: print(f"ℹ️ Havia {len(ids_produtos_oferta_xml)} produtos nas ofertas XML, mas nenhum correspondente foi encontrado ou pode ser colorido em '{PRODUCT_MASTER_FILE_NAME}'.")
            else: print(f"ℹ️ Nenhum produto das ofertas XML para colorir em '{PRODUCT_MASTER_FILE_NAME}'.")
    except Exception as e: print(f"❌ Erro ao salvar ou formatar '{PRODUCT_MASTER_FILE_NAME}': {e}"); traceback.print_exc()


def gerar_relatorio_ofertas_finais(df_ofertas_xml: pd.DataFrame) -> None:
    path_cadastro_base = f'./{PRODUCT_MASTER_FILE_NAME}'
    path_relatorio_final = f'./{FINAL_OFFERS_REPORT_NAME}'
    colunas_finais_excel_ordenadas = ['NOME_PROMOÇÃO', 'SESSÃO', 'ID', 'PRODUTO', 'TIPO', 'PROMOÇÃO']

    try:
        df_cadastro_base = pd.read_excel(path_cadastro_base)
        colunas_cadastro_esperadas = ['ID', 'SESSÃO', 'NOME_SISTEMA', 'NOME_CORRIGIDO']
        if not all(col in df_cadastro_base.columns for col in colunas_cadastro_esperadas):
            missing_cols = [col for col in colunas_cadastro_esperadas if col not in df_cadastro_base.columns]
            print(f"❌ Arquivo '{PRODUCT_MASTER_FILE_NAME}' não contém todas as colunas esperadas. Faltando: {missing_cols}. Não é possível gerar relatório.")
            return
    except FileNotFoundError: print(f"❌ Arquivo de cadastro '{PRODUCT_MASTER_FILE_NAME}' não encontrado."); return
    except Exception as e: print(f"❌ Erro ao carregar '{PRODUCT_MASTER_FILE_NAME}': {e}"); return

    if df_ofertas_xml.empty: print(f"⚠️ DataFrame de ofertas XML vazio. Nada para '{FINAL_OFFERS_REPORT_NAME}'."); return
    required_xml_cols = ['ID_Produto_XML', 'Promoção_XML', 'Preco_Promocao_XML', 'Nome_Produto_XML']
    if not all(col in df_ofertas_xml.columns for col in required_xml_cols):
        print(f"❌ DataFrame de ofertas XML não contém todas as colunas requeridas: {required_xml_cols}."); return

    df_ofertas_xml_copy = df_ofertas_xml[required_xml_cols].copy()
    df_cadastro_base_copy = df_cadastro_base[['ID', 'SESSÃO', 'NOME_CORRIGIDO']].copy()
    df_ofertas_xml_copy['ID_Produto_XML'] = pd.to_numeric(df_ofertas_xml_copy['ID_Produto_XML'], errors='coerce').astype('Int64')
    df_cadastro_base_copy['ID'] = pd.to_numeric(df_cadastro_base_copy['ID'], errors='coerce').astype('Int64')

    df_merged = pd.merge(df_ofertas_xml_copy, df_cadastro_base_copy, left_on='ID_Produto_XML', right_on='ID', how='left')
    df_merged['NOME_CORRIGIDO'] = df_merged['NOME_CORRIGIDO'].fillna(df_merged['Nome_Produto_XML'])
    df_merged['SESSÃO'] = df_merged['SESSÃO'].fillna('SEÇÃO NÃO ESPECIFICADA')
    df_merged['Promoção_XML'] = df_merged['Promoção_XML'].fillna('')
    df_merged['Produto_Base_Agrupamento'] = df_merged['NOME_CORRIGIDO'].apply(lambda x: extract_product_base_name(str(x)))
    df_merged['Tipo_Produto_Variante'] = df_merged.apply(lambda row: extract_product_type_from_full_name(str(row['NOME_CORRIGIDO']), str(row['Produto_Base_Agrupamento'])), axis=1)
    df_merged['Excecao_Agrupamento'] = df_merged['SESSÃO'].isin(SECOES_EXCECAO_AGRUPAMENTO)

    lista_dfs_para_nao_agregados = []
    lista_dfs_para_agregados_reais = []

    df_excecoes_diretas = df_merged[df_merged['Excecao_Agrupamento']].copy()
    if not df_excecoes_diretas.empty:
        lista_dfs_para_nao_agregados.append(
            format_individual_product_for_report(df_excecoes_diretas, colunas_finais_excel_ordenadas)
        )

    df_potenciais = df_merged[~df_merged['Excecao_Agrupamento']].copy()
    if not df_potenciais.empty:
        df_potenciais['Contagem_No_Grupo'] = df_potenciais.groupby(['Produto_Base_Agrupamento', 'Preco_Promocao_XML'])['ID_Produto_XML'].transform('size')

        df_para_agregar = df_potenciais[df_potenciais['Contagem_No_Grupo'] > 1].copy()
        if not df_para_agregar.empty:
            def aggregate_types(series_types):
                valid_types = [str(t).strip() for t in series_types if pd.notna(t) and str(t).strip()]
                return ', '.join(sorted(list(set(valid_types)))) if valid_types else ""
            df_agregados_final = df_para_agregar.groupby(['Produto_Base_Agrupamento', 'Preco_Promocao_XML'], as_index=False).agg(
                TIPO=pd.NamedAgg(column='Tipo_Produto_Variante', aggfunc=aggregate_types),
                NOME_PROMOÇÃO=pd.NamedAgg(column='Promoção_XML', aggfunc='first'),
                ID=pd.NamedAgg(column='ID_Produto_XML', aggfunc='first'),
                SESSÃO=pd.NamedAgg(column='SESSÃO', aggfunc='first')
            ).rename(columns={'Produto_Base_Agrupamento': 'PRODUTO', 'Preco_Promocao_XML': 'PROMOÇÃO'})
            lista_dfs_para_agregados_reais.append(df_agregados_final.reindex(columns=colunas_finais_excel_ordenadas))

        df_unicos_em_grupo = df_potenciais[df_potenciais['Contagem_No_Grupo'] == 1].copy()
        if not df_unicos_em_grupo.empty:
            lista_dfs_para_nao_agregados.append(
                format_individual_product_for_report(df_unicos_em_grupo, colunas_finais_excel_ordenadas)
            )

    df_resultado_agregados = pd.concat(lista_dfs_para_agregados_reais, ignore_index=True) if lista_dfs_para_agregados_reais else pd.DataFrame(columns=colunas_finais_excel_ordenadas)
    df_resultado_nao_agregados = pd.concat(lista_dfs_para_nao_agregados, ignore_index=True) if lista_dfs_para_nao_agregados else pd.DataFrame(columns=colunas_finais_excel_ordenadas)
    df_relatorio_final = pd.concat([df_resultado_agregados, df_resultado_nao_agregados], ignore_index=True)

    for col in ['NOME_PROMOÇÃO', 'SESSÃO', 'PRODUTO', 'TIPO']:
        if col in df_relatorio_final.columns: df_relatorio_final[col] = df_relatorio_final[col].fillna("")
        else: df_relatorio_final[col] = ""
    for col in ['ID', 'PROMOÇÃO']:
        if col not in df_relatorio_final.columns: df_relatorio_final[col] = np.nan

    df_relatorio_final = df_relatorio_final[colunas_finais_excel_ordenadas]
    df_relatorio_final['Chave_Grupo_Secao_Ordenacao'] = df_relatorio_final['SESSÃO'].apply(get_section_sort_key)
    df_ordenado = df_relatorio_final.sort_values(by=['Chave_Grupo_Secao_Ordenacao', 'SESSÃO', 'PRODUTO'], ascending=[True, True, True]).reset_index(drop=True)

    lista_linhas_com_espacos = []
    ultima_chave_grupo_vista = None
    colunas_df_ordenado = df_ordenado.columns.tolist()
    for _, row_data in df_ordenado.iterrows():
        chave_grupo_atual = row_data['Chave_Grupo_Secao_Ordenacao']
        if ultima_chave_grupo_vista is not None and chave_grupo_atual != ultima_chave_grupo_vista:
            lista_linhas_com_espacos.append({col: np.nan for col in colunas_df_ordenado})
        lista_linhas_com_espacos.append(row_data.to_dict())
        ultima_chave_grupo_vista = chave_grupo_atual

    if not lista_linhas_com_espacos: print(f"⚠️ Nenhum dado para gerar o relatório '{FINAL_OFFERS_REPORT_NAME}'."); return
    df_com_espacos = pd.DataFrame(lista_linhas_com_espacos)
    df_output_excel = df_com_espacos.drop(columns=['Chave_Grupo_Secao_Ordenacao'])

    try:
        with pd.ExcelWriter(path_relatorio_final, engine='openpyxl') as writer:
            df_output_excel.to_excel(writer, index=False, sheet_name='Ofertas Finais')
            ws = writer.sheets['Ofertas Finais']
            ws.sheet_view.showGridLines = False

            excel_column_names = list(df_output_excel.columns)
            thin_border_side = Side(style='thin', color="BFBFBF") # Cinza claro para bordas internas
            header_bottom_border_side = Side(style='medium', color="000000") # Borda preta mais grossa abaixo do cabeçalho

            default_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=header_bottom_border_side)


            # Aplicar formatação de coluna e cabeçalho
            for col_idx_excel, column_name in enumerate(excel_column_names, 1):
                base_config = COLUMNS_FORMATTING_CONFIG.get('_DEFAULT_', {}).copy()
                specific_config = COLUMNS_FORMATTING_CONFIG.get(column_name, {}).copy()
                col_config = {**base_config, **specific_config}

                header_cell = ws.cell(row=1, column=col_idx_excel)
                header_cell.font = Font(name=col_config.get('header_font_name'),
                                          sz=col_config.get('header_font_size'),
                                          bold=col_config.get('header_font_bold'),
                                          italic=col_config.get('header_font_italic'),
                                          color=col_config.get('header_font_color'))
                if col_config.get('header_fill_color'):
                    header_cell.fill = PatternFill(start_color=col_config.get('header_fill_color'),
                                                   end_color=col_config.get('header_fill_color'),
                                                   fill_type="solid")
                header_cell.alignment = Alignment(horizontal=col_config.get('header_alignment_horizontal'),
                                                  vertical=col_config.get('header_alignment_vertical'),
                                                  wrap_text=True) # Quebra de texto para cabeçalhos
                header_cell.border = header_border # Borda específica para cabeçalho

            # Aplicar formatação de dados e bordas (linha por linha, pulando espaçadores)
            for row_idx_excel in range(2, ws.max_row + 1):
                is_spacer_row = all(
                    (ws.cell(row=row_idx_excel, column=c_idx_check).value is None or
                     str(ws.cell(row=row_idx_excel, column=c_idx_check).value).strip() == '' or
                     pd.isna(ws.cell(row=row_idx_excel, column=c_idx_check).value)
                    ) for c_idx_check in range(1, ws.max_column + 1)
                )
                if is_spacer_row: continue

                for col_idx_excel, column_name in enumerate(excel_column_names, 1):
                    cell = ws.cell(row=row_idx_excel, column=col_idx_excel)

                    base_config = COLUMNS_FORMATTING_CONFIG.get('_DEFAULT_', {}).copy()
                    specific_config = COLUMNS_FORMATTING_CONFIG.get(column_name, {}).copy()
                    col_config = {**base_config, **specific_config}

                    cell.font = Font(name=col_config.get('font_name'),
                                      sz=col_config.get('font_size'),
                                      bold=col_config.get('font_bold'),
                                      italic=col_config.get('font_italic'),
                                      color=col_config.get('font_color'))
                    cell.alignment = Alignment(horizontal=col_config.get('alignment_horizontal'),
                                               vertical=col_config.get('alignment_vertical'))

                    if cell.value is not None and col_config.get('number_format'):
                        # Se for 0.0 e o formato for contábil, garantir 0.00
                        if col_config.get('number_format') == 'R$ #,##0.00' and cell.value == 0.0:
                             cell.value = 0.00 # Garante que apareça como 0.00 e não -
                        cell.number_format = col_config.get('number_format')

                    cell.border = default_border

            # Ajuste de largura das colunas
            for col_idx_excel, column_name in enumerate(excel_column_names, 1):
                base_config = COLUMNS_FORMATTING_CONFIG.get('_DEFAULT_', {}).copy()
                specific_config = COLUMNS_FORMATTING_CONFIG.get(column_name, {}).copy()
                col_config = {**base_config, **specific_config}

                column_letter = get_column_letter(col_idx_excel)
                max_len = 0

                header_text = str(ws.cell(row=1, column=col_idx_excel).value)
                # Aproximação da largura do cabeçalho (pode precisar de ajuste se fontes muito diferentes)
                max_len = max(max_len, len(header_text) * (col_config.get('header_font_size', 11) / 10.0) * 1.0)


                for r_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r_idx, column=col_idx_excel)
                    if cell.value is not None:
                        # Obter configuração da coluna para esta célula de dados
                        data_font_size = col_config.get('font_size', 10)

                        if col_config.get('number_format') == 'R$ #,##0.00' and isinstance(cell.value, (int, float)):
                            text_val = f"R$ {cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        elif col_config.get('number_format') == '0' and isinstance(cell.value, (int, float)):
                            text_val = f"{int(cell.value)}"
                        else:
                            text_val = str(cell.value)

                        # Fator de ajuste aproximado para tamanho da fonte
                        current_len = len(text_val) * (data_font_size / 10.0) * 1.1 # Fator 1.1 para respiro
                        max_len = max(max_len, current_len)

                # Largura mínima de 8, máxima de 50. Adiciona um pouco de padding.
                adjusted_width = max(8, min(max_len + 2, 50))
                ws.column_dimensions[column_letter].width = adjusted_width

        print(f"✅ Relatório '{FINAL_OFFERS_REPORT_NAME}' criado com {len(df_ordenado.dropna(subset=['PRODUTO'], how='all'))} ofertas e formatação avançada.")
    except Exception as e: print(f"❌ Erro ao salvar/formatar Excel '{FINAL_OFFERS_REPORT_NAME}': {e}"); traceback.print_exc()

# FUNÇÃO PRINCIPAL
# ... (igual ao anterior)
def main() -> None:
    print("🚀 Iniciando processamento de ofertas...")
    if not os.path.isdir(INPUT_XML_DIR): print(f"❌ ERRO: Diretório '{INPUT_XML_DIR}' não encontrado."); return
    arquivos_xml_originais = listar_arquivos_xml(INPUT_XML_DIR)
    if not arquivos_xml_originais: print(f"ℹ️ Nenhum arquivo XML encontrado em '{INPUT_XML_DIR}'.")
    else: print(f"🔄 Normalizando nomes de {len(arquivos_xml_originais)} arquivos XML..."); renomear_arquivos_em_lote(arquivos_xml_originais, INPUT_XML_DIR)

    arquivos_xml_para_processar = [f for f in os.listdir(INPUT_XML_DIR) if f.endswith(XML_EXTENSION) and os.path.isfile(os.path.join(INPUT_XML_DIR, f))]
    if not arquivos_xml_para_processar: print(f"ℹ️ Nenhum arquivo XML para processar em '{INPUT_XML_DIR}' após normalização.")

    print(f"📄 Consolidando dados de {len(arquivos_xml_para_processar)} arquivos XML...")
    df_ofertas_consolidadas_xml = consolidar_dados_de_xmls(arquivos_xml_para_processar)
    if df_ofertas_consolidadas_xml.empty: print("ℹ️ Nenhum dado de produto consolidado dos XMLs.")

    print(f"💾 Atualizando cadastro: './{PRODUCT_MASTER_FILE_NAME}'...")
    atualizar_cadastro_produtos_base(df_ofertas_consolidadas_xml)

    if df_ofertas_consolidadas_xml.empty: print("🏁 Processamento concluído (sem dados XML para relatório de ofertas)."); return

    print(f"📊 Gerando relatório final: './{FINAL_OFFERS_REPORT_NAME}'...")
    gerar_relatorio_ofertas_finais(df_ofertas_consolidadas_xml)
    print("🏁 Processamento concluído.")

if __name__ == "__main__":
    main()