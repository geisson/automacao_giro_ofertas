# -*- coding: utf-8 -*-
"""
Script para processamento de arquivos XML de ofertas, refatorado com paradigma funcional.
Dividido em seções bem definidas com funções puras sempre que possível.
Modificado para incluir agrupamento condicional de produtos, formatação de Excel
avançada (contábil, bordas) e ordenação específica de seções com linhas em branco.
"""

# ------------------------------------------
# DEPENDÊNCIAS
# ------------------------------------------
import pandas as pd
import os
import glob
import xmltodict
from typing import List, Dict, Any
import re
import numpy as np
from openpyxl.styles import Border, Side, Alignment, Font # Para formatação Excel
from openpyxl.utils import get_column_letter # Para ajuste de largura de coluna
import traceback # Para melhor depuração de erros

# ------------------------------------------
# CONSTANTES
# ------------------------------------------
INPUT_DIR = './dados_brutos/'
OUTPUT_DIR = './'
XML_EXTENSION = '.xml'
CORRECTED_FILE = 'dados_produtos_corrigidos.xlsx'
CONSOLIDATED_FILE = 'ofertas_consolidadas.xlsx'
CORRECTED_OFFERS_FILE = 'ofertas_corrigidas.xlsx'

# ------------------------------------------
# FUNÇÕES PARA TRATAMENTO DE TEXTO
# ------------------------------------------

def remover_acentuacao(texto: str) -> str:
    mapeamento_acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u', 'ã': 'a', 'õ': 'o', 'ä': 'a', 'ë': 'e', 'ï': 'i',
        'ö': 'o', 'ü': 'u', 'ç': 'c', 'ñ': 'n', 'ÿ': 'y', 'ý': 'y', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O',
        'Ú': 'U', 'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U', 'Â': 'A', 'Ê': 'E', 'Î': 'I', 'Ô': 'O',
        'Û': 'U', 'Ã': 'A', 'Õ': 'O', 'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U', 'Ç': 'C', 'Ñ': 'N',
        'Ÿ': 'Y', 'Ý': 'Y'
    }
    return ''.join(mapeamento_acentos.get(c, c) for c in texto)

def normalizar_texto(texto: str) -> str:
    texto_sem_acentos = remover_acentuacao(texto)
    return texto_sem_acentos.lower()

# ------------------------------------------
# FUNÇÕES PARA MANIPULAÇÃO DE ARQUIVOS
# ------------------------------------------

def listar_arquivos_xml(diretorio: str) -> List[str]:
    return glob.glob(os.path.join(diretorio, f'*{XML_EXTENSION}'))

def renomear_arquivo(arquivo_antigo: str, diretorio: str) -> None:
    nome_arquivo = os.path.basename(arquivo_antigo)
    novo_nome = normalizar_texto(nome_arquivo)
    novo_caminho = os.path.join(diretorio, novo_nome)
    try:
        if arquivo_antigo != novo_caminho:
            os.rename(arquivo_antigo, novo_caminho)
    except FileExistsError:
        print(f"⚠️  Arquivo '{novo_caminho}' já existe. Não foi renomeado '{arquivo_antigo}'.")
    except Exception as e:
        print(f"❌ Erro ao renomear '{arquivo_antigo}' para '{novo_caminho}': {e}")

def renomear_arquivos_em_lote(arquivos: List[str], diretorio: str) -> None:
    for arquivo in arquivos:
        renomear_arquivo(arquivo, diretorio)

# ------------------------------------------
# FUNÇÕES PARA PROCESSAMENTO DE XML
# ------------------------------------------

def parse_produto(produto: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "Promoção": produto.get('descrpromocao', ''),
        "ID Produto": int(produto.get('idsubproduto', 0)),
        "Produto": produto.get('descrresproduto', ''),
        "Preço Promoção": float(produto.get('precopromocao', 0.0))
    }

def processar_xml(nome_arquivo: str) -> List[Dict[str, Any]]:
    caminho_arquivo = os.path.join(INPUT_DIR, nome_arquivo)
    try:
        with open(caminho_arquivo, 'rb') as arquivo_xml:
            dados_xml = xmltodict.parse(arquivo_xml)
            if 'temporario_846' not in dados_xml:
                print(f"⚠️ Estrutura 'temporario_846' não encontrada em {nome_arquivo}.")
                return []
            temporario_846 = dados_xml['temporario_846']
            produtos_data = temporario_846.get('temporario_846_row', [])
            if isinstance(produtos_data, dict):
                produtos_data = [produtos_data]
            if not produtos_data:
                return []
            return [parse_produto(p) for p in produtos_data if isinstance(p, dict)]
    except FileNotFoundError:
        print(f"❌ Arquivo XML não encontrado: {caminho_arquivo}")
        return []
    except Exception as e:
        print(f"❌ Erro ao processar XML {nome_arquivo}: {e}")
        return []

# ------------------------------------------
# FUNÇÕES PARA CONSOLIDAÇÃO DE DADOS
# ------------------------------------------

def consolidar_produtos(lista_arquivos: List[str]) -> List[Dict[str, Any]]:
    todos_produtos = []
    for arquivo in lista_arquivos:
        produtos = processar_xml(arquivo)
        if produtos:
            todos_produtos.extend(produtos)
            print(f"✅ {arquivo}: {len(produtos)} produtos processados.")
    return todos_produtos

def gerar_planilha_ofertas(produtos: List[Dict[str, Any]]) -> None:
    if not produtos:
        print(f"⚠️ Nenhum dado de produto foi processado para gerar '{CONSOLIDATED_FILE}'.")
        return
    try:
        df = pd.DataFrame(produtos)
        df.to_excel(CONSOLIDATED_FILE, index=False)
        print(f"🎉 Excel '{CONSOLIDATED_FILE}' gerado com {len(df)} produtos!")
    except Exception as e:
        print(f"❌ Erro ao gerar planilha '{CONSOLIDATED_FILE}': {e}")

# ------------------------------------------
# FUNÇÕES AUXILIARES PARA INTEGRAÇÃO DE DADOS
# ------------------------------------------

def extract_product_base_name(product_name: str) -> str:
    if not isinstance(product_name, str): return ""
    unidades = [
        "KILOGRAMAS", "KILOGRAMA", "QUILOGRAMAS", "QUILOGRAMA", "LITROS", "LITRO", "UNIDADES", "UNIDADE",
        "PACOTES", "PACOTE", "CAIXAS", "CAIXA", "FARDOS", "FARDO", "GARRAFAS", "GARRAFA", "LATAS", "LATA",
        "POTES", "POTE", "ROLOS", "ROLO", "SACHES", "SACHE", "TUBOS", "TUBO", "VIDROS", "VIDRO", "SACOLAS",
        "SACOLA", "BLISTERS", "BLISTER", "CARTELAS", "CARTELA", "DISPLAYS", "DISPLAY", "GALÕES", "GALÃO",
        "REFILs", "REFIL", "TABLETES", "TABLETE", "PEÇAS", "PEÇA", "FOLHAS", "FOLHA", "FLACONETES",
        "FLACONETE", "BISNAGAS", "BISNAGA", "AMPOLAS", "AMPOLA", "DRAGEAS", "CAPSULAS", "CAPSULA",
        "COMPRIMIDOS", "COMPRIMIDO", "ENVELOPES", "ENVELOPE", "KG", "GR", "G", "MG", "L", "ML", "CM", "MM",
        "M", "UN", "UND", "UNID", "PC", "PÇ", "PCT", "CX", "FD", "GF", "LT", "PT", "RL", "SC", "TB", "VD",
        "BL", "CT", "DP", "GL", "RF", "TBL", "FL", "CAPS", "COMP", "ENV"
    ]
    unidades_regex_part = "|".join(map(re.escape, sorted(unidades, key=len, reverse=True)))
    regex_str = r"^(.*?)(\s?)(\d+[\.,]?\d*)(\s*)(" + unidades_regex_part + r")\b.*$"
    match = re.match(regex_str, product_name, re.IGNORECASE)
    if match:
        nome_antes, _, quantidade, espaco_unidade, unidade_texto = match.groups()[:5]
        nome_antes = nome_antes.strip()
        base_nome = f"{nome_antes} {quantidade}{espaco_unidade}{unidade_texto}" if nome_antes else f"{quantidade}{espaco_unidade}{unidade_texto}"
        return base_nome.strip().replace("  ", " ")
    return product_name

def get_section_group_key(section_name: Any) -> str:
    if not isinstance(section_name, str) or pd.isna(section_name): return "SEM SEÇÃO DEFINIDA"
    if section_name.startswith("#01 MERCEARIA - #") and "ALTO GIRO" in section_name:
        return "#01 MERCEARIA - ALTO GIRO"
    return section_name.split(" - ", 1)[0]

# ------------------------------------------
# FUNÇÕES PARA INTEGRAÇÃO DE DADOS (PRINCIPAIS)
# ------------------------------------------

def adicionar_novos_produtos() -> None:
    try:
        df_corrigidos = pd.read_excel(CORRECTED_FILE)
    except FileNotFoundError:
        print(f"⚠️  Arquivo '{CORRECTED_FILE}' não encontrado. Será criado.")
        df_corrigidos = pd.DataFrame(columns=['ID Produto', 'Produto', 'Preço Promoção', 'Seção', 'Produto Corrigido'])
    except Exception as e:
        print(f"❌ Erro ao carregar '{CORRECTED_FILE}': {e}"); return

    try:
        df_ofertas = pd.read_excel(CONSOLIDATED_FILE)
    except FileNotFoundError:
        print(f"ℹ️ Arquivo '{CONSOLIDATED_FILE}' não encontrado. Nada a adicionar."); return
    except Exception as e:
        print(f"❌ Erro ao carregar '{CONSOLIDATED_FILE}': {e}"); return

    if df_ofertas.empty: print(f"ℹ️ '{CONSOLIDATED_FILE}' vazio. Nada a adicionar."); return
    if 'ID Produto' not in df_ofertas.columns: print(f"❌ 'ID Produto' ausente em '{CONSOLIDATED_FILE}'."); return
    if 'ID Produto' not in df_corrigidos.columns and not df_corrigidos.empty:
         print(f"❌ 'ID Produto' ausente em '{CORRECTED_FILE}'. Não é possível verificar duplicatas.")

    ids_existentes = set(df_corrigidos['ID Produto'].unique()) if 'ID Produto' in df_corrigidos.columns else set()
    novos_produtos = df_ofertas[~df_ofertas['ID Produto'].isin(ids_existentes)]

    if novos_produtos.empty: print(f"ℹ️ Nenhum produto novo de '{CONSOLIDATED_FILE}' para '{CORRECTED_FILE}'."); return

    colunas_para_novos = ['ID Produto', 'Produto', 'Preço Promoção']
    novos_produtos_filtrados = novos_produtos[colunas_para_novos].copy()

    for col in df_corrigidos.columns:
        if col not in novos_produtos_filtrados.columns:
            if col == 'Seção': novos_produtos_filtrados[col] = 'SEÇÃO NÃO ESPECIFICADA'
            elif col == 'Produto Corrigido':
                novos_produtos_filtrados[col] = novos_produtos_filtrados['Produto'] if 'Produto' in novos_produtos_filtrados else 'PRODUTO NÃO ESPECIFICADO'
            else: novos_produtos_filtrados[col] = np.nan

    novos_produtos_alinhados = novos_produtos_filtrados.reindex(columns=df_corrigidos.columns, fill_value=np.nan)
    df_final = pd.concat([df_corrigidos, novos_produtos_alinhados], ignore_index=True)
    df_final.drop_duplicates(subset=['ID Produto'], keep='first', inplace=True)

    try:
        df_final.to_excel(CORRECTED_FILE, index=False)
        print(f"✅ '{CORRECTED_FILE}' atualizado com {len(novos_produtos_alinhados)} novos produtos.")
    except Exception as e:
        print(f"❌ Erro ao salvar '{CORRECTED_FILE}': {e}")

def gerar_ofertas_corrigidas() -> None:
    try:
        df_corrigidos = pd.read_excel(CORRECTED_FILE)
    except FileNotFoundError: print(f"❌ '{CORRECTED_FILE}' não encontrado."); return
    except Exception as e: print(f"❌ Erro ao carregar '{CORRECTED_FILE}': {e}"); return

    try:
        df_ofertas = pd.read_excel(CONSOLIDATED_FILE)
    except FileNotFoundError: print(f"❌ '{CONSOLIDATED_FILE}' não encontrado."); return
    except Exception as e: print(f"❌ Erro ao carregar '{CONSOLIDATED_FILE}': {e}"); return

    if df_ofertas.empty: print(f"⚠️ '{CONSOLIDATED_FILE}' vazio."); return
    if df_corrigidos.empty: print(f"⚠️ '{CORRECTED_FILE}' vazio. 'Seção'/'Produto Corrigido' podem ter padrões.")

    df_merged = df_ofertas[['ID Produto', 'Promoção', 'Preço Promoção']].merge(
        df_corrigidos[['ID Produto', 'Seção', 'Produto Corrigido']], on='ID Produto', how='left'
    )

    df_merged['Produto Corrigido'] = df_merged['Produto Corrigido'].fillna(df_merged['ID Produto'].astype(str) + '_NOME_ORIGINAL_FALTANTE')
    df_merged['Seção'] = df_merged['Seção'].fillna('SEÇÃO NÃO ESPECIFICADA')
    df_merged['Promoção_Descricao_Textual'] = df_merged['Promoção'].fillna('')
    df_merged['Produto_Base_Extraido'] = df_merged['Produto Corrigido'].apply(extract_product_base_name)
    group_cols_for_counting = ['Produto_Base_Extraido', 'Preço Promoção']
    df_merged['Contagem_Grupo_Produto_Preco'] = df_merged.groupby(group_cols_for_counting)['ID Produto'].transform('size')
    df_merged['Nome_Produto_Final_Agrupamento'] = np.where(
        df_merged['Contagem_Grupo_Produto_Preco'] > 1, df_merged['Produto_Base_Extraido'], df_merged['Produto Corrigido']
    )

    df_aggregated = df_merged.groupby(
        ['Nome_Produto_Final_Agrupamento', 'Preço Promoção'], as_index=False
    ).agg(
        Promo_Desc_Agg=pd.NamedAgg(column='Promoção_Descricao_Textual', aggfunc='first'),
        ID_Produto_Agg=pd.NamedAgg(column='ID Produto', aggfunc='first'),
        Seção_Agg=pd.NamedAgg(column='Seção', aggfunc='first')
    )

    df_final_renamed = df_aggregated.rename(columns={
        'Nome_Produto_Final_Agrupamento': 'PRODUTO_Final', 'Preço Promoção': 'PROMOÇÃO_Valor_Num',
        'Promo_Desc_Agg': 'PROMOÇÃO_Descricao_Txt', 'ID_Produto_Agg': 'ID_Final', 'Seção_Agg': 'Seção_Final'
    })

    colunas_ordem = ['PROMOÇÃO_Descricao_Txt', 'Seção_Final', 'ID_Final', 'PRODUTO_Final', 'PROMOÇÃO_Valor_Num']
    df_para_ordenar = df_final_renamed[colunas_ordem].copy()
    # Define as colunas como você especificou: ['Promoção', 'Seção', 'ID', 'PRODUTO', 'PROMOÇÃO']
    df_para_ordenar.columns = ['Promoção', 'Seção', 'ID', 'PRODUTO', 'PROMOÇÃO']
    df_para_ordenar['Chave_Grupo_Secao_Ordenacao'] = df_para_ordenar['Seção'].apply(get_section_group_key)

    df_ordenado = df_para_ordenar.sort_values(
        by=['Chave_Grupo_Secao_Ordenacao', 'Seção', 'PRODUTO'], ascending=[True, True, True]
    ).reset_index(drop=True)

    lista_linhas_com_espacos = []
    ultima_chave_grupo_vista = None
    colunas_df_ordenado = df_ordenado.columns
    for _, row_data in df_ordenado.iterrows():
        chave_grupo_atual = row_data['Chave_Grupo_Secao_Ordenacao']
        if ultima_chave_grupo_vista is not None and chave_grupo_atual != ultima_chave_grupo_vista:
            lista_linhas_com_espacos.append({col: np.nan for col in colunas_df_ordenado})
        lista_linhas_com_espacos.append(row_data.to_dict())
        ultima_chave_grupo_vista = chave_grupo_atual

    if not lista_linhas_com_espacos: print(f"⚠️ Nenhum dado para '{CORRECTED_OFFERS_FILE}'."); return

    df_com_espacos = pd.DataFrame(lista_linhas_com_espacos)
    df_output_final = df_com_espacos.drop(columns=['Chave_Grupo_Secao_Ordenacao'])

    try:
        with pd.ExcelWriter(CORRECTED_OFFERS_FILE, engine='openpyxl') as writer:
            df_output_final.to_excel(writer, index=False, sheet_name='Ofertas Corrigidas')
            ws = writer.sheets['Ofertas Corrigidas']

            thin_side = Side(style='thin')
            border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            formato_contabil = 'R$ #,##0.00'

            # Correção: Identificar a coluna de valor numérico "PROMOÇÃO" (a última)
            col_idx_valor_promo_excel = -1
            if df_output_final.columns[-1] == 'PROMOÇÃO':
                # A última coluna é de índice len(df_output_final.columns) - 1
                # Convertendo para 1-based para openpyxl:
                col_idx_valor_promo_excel = len(df_output_final.columns)
            else:
                print(f"⚠️ Aviso: A última coluna não é 'PROMOÇÃO'. É '{df_output_final.columns[-1]}'. Não formatando como contábil.")


            for row_idx in range(1, ws.max_row + 1):
                is_spacer_row_in_excel = True
                for c_idx in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=row_idx, column=c_idx).value
                    if cell_val is not None and cell_val != "":
                        is_spacer_row_in_excel = False
                        break

                if is_spacer_row_in_excel:
                    continue

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border_style
                    if row_idx > 1 and col_idx == col_idx_valor_promo_excel and cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = formato_contabil
                        else:
                            try:
                                if isinstance(cell.value, str):
                                    cleaned_value = cell.value.replace("R$", "").strip()
                                    parts = cleaned_value.split(',')
                                    if len(parts) == 2:
                                        integer_part = parts[0].replace(".", "")
                                        cleaned_value = f"{integer_part}.{parts[1]}"
                                    elif cleaned_value.count('.') > 1: # Ex: 1.234.567 -> 1234567 (sem vírgula)
                                        cleaned_value = cleaned_value.replace(".","")
                                    # Se for apenas "1234" ou "1234.56" (já com ponto decimal), não precisa de muita alteração
                                    # Esta parte pode precisar de mais robustez dependendo dos formatos de entrada.
                                    numeric_value = float(cleaned_value)
                                    cell.value = numeric_value
                                    cell.number_format = formato_contabil
                            except ValueError:
                                print(f"⚠️ Não convertível: '{cell.value}' L{row_idx},C{col_idx}.")

            for col_openpyxl_idx in range(1, ws.max_column + 1):
                column_letter_id = get_column_letter(col_openpyxl_idx)
                max_len = 0
                for cell_tuple in ws.iter_cols(min_col=col_openpyxl_idx, max_col=col_openpyxl_idx):
                    for cell in cell_tuple:
                        if cell.value is not None:
                            is_header = cell.row == 1
                            is_formatted_promo_col = (col_openpyxl_idx == col_idx_valor_promo_excel and
                                                      not is_header and
                                                      cell.number_format == formato_contabil and
                                                      isinstance(cell.value, (int, float)))

                            if is_formatted_promo_col:
                                text_val = f"R$ {cell.value:,.2f}"
                                max_len = max(max_len, len(text_val))
                            else:
                                max_len = max(max_len, len(str(cell.value)))
                adjusted_width = min(max_len + 3, 50)
                ws.column_dimensions[column_letter_id].width = adjusted_width if adjusted_width > 5 else 10

        num_ofertas_reais = len(df_ordenado.dropna(subset=['PRODUTO'], how='all'))
        print(f"✅ Tabela '{CORRECTED_OFFERS_FILE}' criada com {num_ofertas_reais} ofertas.")
    except Exception as e:
        print(f"❌ Erro ao salvar/formatar Excel '{CORRECTED_OFFERS_FILE}': {e}")
        traceback.print_exc()

# ------------------------------------------
# FUNÇÃO PRINCIPAL
# ------------------------------------------

def main() -> None:
    os.makedirs(INPUT_DIR, exist_ok=True)
    arquivos_xml_originais = listar_arquivos_xml(INPUT_DIR)
    if not arquivos_xml_originais: print(f"ℹ️ Nenhum XML em '{INPUT_DIR}'.")
    else:
        print(f"Normalizando {len(arquivos_xml_originais)} arquivos XML...")
        renomear_arquivos_em_lote(arquivos_xml_originais, INPUT_DIR)

    arquivos_xml_para_processar = [f for f in os.listdir(INPUT_DIR) if f.endswith(XML_EXTENSION) and os.path.isfile(os.path.join(INPUT_DIR, f))]
    if not arquivos_xml_para_processar: print(f"ℹ️ Nenhum XML para processar em '{INPUT_DIR}'.")

    todos_produtos = consolidar_produtos(arquivos_xml_para_processar)
    gerar_planilha_ofertas(todos_produtos)
    adicionar_novos_produtos()
    gerar_ofertas_corrigidas()
    print("🏁 Processamento concluído.")

if __name__ == "__main__":
    main()